import os
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from io import BytesIO
import smtplib
import ssl
import certifi
import re
from email.mime.text import MIMEText
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# KYC Module Enabled - Deployed March 2026

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from flask_sqlalchemy import SQLAlchemy
from dotenv import load_dotenv
from sqlalchemy import or_, and_, text, func
from flask_login import (
    LoginManager, login_user, logout_user, login_required, current_user, UserMixin
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix

# Load environment variables from .env if present
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', 'dev-secret')

# Normalize DATABASE_URL for SQLAlchemy (supports postgres:// -> postgresql://)
db_url = os.getenv('DATABASE_URL', 'sqlite:///escalations.db')
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Respect reverse proxy headers (scheme, host, prefix) on PaaS / behind Nginx
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

TEAMS = [
    'International Operations',
    'Domestic Operations',
]

PRIORITIES = ['Low', 'Medium', 'High', 'Urgent']

STATUSES = ['Open', 'In Progress', 'Resolved']


class Ticket(db.Model):
    __tablename__ = 'tickets'
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)

    source = db.Column(db.String(50))  # Gmail, LinkedIn, WhatsApp, Other
    customer_name = db.Column(db.String(120))
    contact = db.Column(db.String(120))  # email or phone
    booking_id = db.Column(db.String(80))

    team = db.Column(db.String(50), nullable=False)
    priority = db.Column(db.String(20), default='Medium', nullable=False)
    status = db.Column(db.String(20), default='Open', nullable=False)

    assigned_to = db.Column(db.String(120))
    # Lead assignee - a single user responsible for the ticket
    lead_assignee_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    # Comma-separated notification emails for this ticket
    notify_emails = db.Column(db.Text)

    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    # Track last update time; ensures NOT NULL insert and auto-update on changes
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    resolved_at = db.Column(db.DateTime)
    resolved_by = db.Column(db.String(120))
    # Soft delete
    deleted_at = db.Column(db.DateTime, nullable=True)
    # Relationships
    comments = db.relationship('Comment', backref='ticket', lazy=True, cascade='all, delete-orphan')
    assignees = db.relationship('User', secondary='ticket_assignees', backref='assigned_tickets')
    lead_assignee = db.relationship('User', foreign_keys=[lead_assignee_id], lazy=True)


# Association table for many-to-many Ticket<->User
ticket_assignees = db.Table(
    'ticket_assignees',
    db.Column('ticket_id', db.Integer, db.ForeignKey('tickets.id'), primary_key=True),
    db.Column('user_id', db.Integer, db.ForeignKey('users.id'), primary_key=True),
)


class User(db.Model, UserMixin):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), default='agent', nullable=False)  # 'admin' or 'agent'
    department = db.Column(db.String(80))  # International Operations / Domestic Operations

    def set_password(self, password: str):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        # Primary: verify using Werkzeug hash (pbkdf2/scrypt). Handle invalid/legacy hashes safely.
        try:
            if self.password_hash and check_password_hash(self.password_hash, password):
                return True
        except ValueError:
            # e.g., "Invalid hash method ''" from legacy/empty hashes
            pass

        # Fallback: if DB stored plaintext historically, allow once and upgrade hash
        try:
            if self.password_hash == password and password:
                # Upgrade to a secure hash transparently
                self.set_password(password)
                try:
                    db.session.commit()
                except Exception as e:
                    print(f"[warn] failed to upgrade password hash for user {self.id}: {e}")
                return True
        except Exception:
            pass

        return False


class Founder(db.Model):
    __tablename__ = 'founders'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), nullable=False)


class Comment(db.Model):
    __tablename__ = 'comments'
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False)
    author = db.Column(db.String(120))
    body = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class Notification(db.Model):
    __tablename__ = 'notifications'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    type = db.Column(db.String(50))
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'))
    message = db.Column(db.Text)
    actor_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    is_read = db.Column(db.Boolean, default=False, nullable=False)


class AppMeta(db.Model):
    __tablename__ = 'app_meta'
    key = db.Column(db.String(50), primary_key=True)
    value = db.Column(db.String(200))


# ==================== KYC MODULE ====================

class KYCCustomer(db.Model):
    __tablename__ = 'kyc_customers'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    trip_type = db.Column(db.String(20), nullable=False, default='domestic')  # 'domestic' or 'international'
    trip_name = db.Column(db.String(200))  # Trip name for grouping participants
    trip_date = db.Column(db.Date)  # Trip start date for deadline tracking
    booking_id = db.Column(db.String(80))
    requires_dl = db.Column(db.Boolean, default=True, nullable=False)  # Whether DL upload is required
    deleted_at = db.Column(db.DateTime, nullable=True)  # Soft delete
    created_by_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    group_link_id = db.Column(db.Integer, db.ForeignKey('kyc_group_links.id'), nullable=True)
    
    # KYC completion tracking
    kyc_submitted = db.Column(db.Boolean, default=False, nullable=False)
    kyc_submitted_at = db.Column(db.DateTime)
    
    # Indemnity signing tracking
    indemnity_signed = db.Column(db.Boolean, default=False, nullable=False)
    indemnity_signed_at = db.Column(db.DateTime)
    
    # Unique tokens for external access
    kyc_token = db.Column(db.String(64), unique=True)
    indemnity_token = db.Column(db.String(64), unique=True)

    # Phase 3: NPS (Net Promoter Score) feedback. All nullable so existing
    # rows remain untouched. Rating is 1..5 (1 = angry, 5 = love).
    nps_rating = db.Column(db.Integer)
    nps_feedback = db.Column(db.Text)
    nps_submitted_at = db.Column(db.DateTime)
    
    created_by = db.relationship('User', foreign_keys=[created_by_id], lazy=True)
    submission = db.relationship('KYCSubmission', backref='customer', uselist=False, lazy=True)

    # ---------------------------------------------------------------
    # Phase 6c: DERIVED STATUS PROPERTIES (read-only, source-of-truth)
    # ---------------------------------------------------------------
    # The denormalised booleans `kyc_submitted` / `indemnity_signed` have
    # historically drifted from reality (transient MVCC issues, pre-Phase-1
    # rows, concurrent group-booking submits). These properties return the
    # OR of the raw flag and the presence of the underlying source row, so
    # the UI always shows the truth even if the flag hasn't caught up.
    #
    # IMPORTANT: these do NOT write anything. A separate one-time backfill
    # in ensure_schema() heals the raw flags so the database itself is
    # consistent. Derivation is the belt; backfill is the suspenders.
    #
    # Cost model:
    #   * `effective_kyc_submitted` uses the existing `submission` one-to-one
    #     relationship → zero extra queries when eager-loaded.
    #   * `effective_indemnity_signed` uses the `indemnity_requests` backref
    #     (see IndemnityRequest.customer). In list views we eager-load this
    #     to avoid N+1; in single-customer views the extra query is trivial.
    @property
    def effective_kyc_submitted(self):
        if self.kyc_submitted:
            return True
        # `submission` is a one-to-one backref. Accessing it when lazy-loaded
        # triggers at most one SELECT; when joinedload'd this is free.
        try:
            return self.submission is not None
        except Exception:
            return False

    @property
    def effective_indemnity_signed(self):
        if self.indemnity_signed:
            return True
        try:
            reqs = getattr(self, 'indemnity_requests', None) or []
            return any(r.signed_at is not None for r in reqs)
        except Exception:
            return False

    def get_days_until_trip(self):
        """Calculate days remaining until trip date."""
        if not self.trip_date:
            return None
        from datetime import date
        today = date.today()
        if isinstance(self.trip_date, datetime):
            trip_date = self.trip_date.date()
        else:
            trip_date = self.trip_date
        return (trip_date - today).days
    
    def is_trip_started(self):
        """Check if trip date has passed."""
        if not self.trip_date:
            return False
        from datetime import date
        today = date.today()
        if isinstance(self.trip_date, datetime):
            trip_date = self.trip_date.date()
        else:
            trip_date = self.trip_date
        return trip_date < today
    
    def needs_urgent_attention(self):
        """Check if customer needs urgent attention (trip within 7 days and KYC not done).
        Phase 6c: uses derived status so stuck-flag rows are not mis-flagged
        as urgent when they have in fact completed KYC/indemnity."""
        days = self.get_days_until_trip()
        if days is None:
            return False
        return days <= 7 and (
            not self.effective_kyc_submitted or not self.effective_indemnity_signed
        )

    def count_previous_trips(self):
        """Phase 4: Count other (non-deleted) KYC customer records that
        share this customer's email OR phone, excluding self.

        Strictly read-only — executes a single SELECT COUNT(*) and mutates
        nothing. Returns 0 safely if both email and phone are missing.

        Matching rules:
          - Email: case-insensitive, whitespace-trimmed equality.
          - Phone: whitespace-trimmed equality (numbers are stored as-is
            and we deliberately do NOT strip formatting so we don't
            conflate "+91 98…" with "98…"; that would require a data
            migration we are not doing here).
          - Soft-deleted customers are excluded.
        """
        email = (self.email or '').strip().lower()
        phone = (self.phone or '').strip()
        if not email and not phone:
            return 0
        conditions = []
        if email:
            conditions.append(func.lower(func.trim(KYCCustomer.email)) == email)
        if phone:
            conditions.append(func.trim(KYCCustomer.phone) == phone)
        try:
            return (
                KYCCustomer.query
                .filter(KYCCustomer.id != self.id)
                .filter(KYCCustomer.deleted_at.is_(None))
                .filter(or_(*conditions))
                .count()
            )
        except Exception as e:
            # Never fail a page render over a counter.
            print(f"[warn] count_previous_trips failed for customer {self.id}: {e}")
            return 0


class KYCGroupLink(db.Model):
    __tablename__ = 'kyc_group_links'
    id = db.Column(db.Integer, primary_key=True)
    token = db.Column(db.String(64), unique=True, nullable=False)
    booking_id = db.Column(db.String(80))
    trip_name = db.Column(db.String(200))
    trip_type = db.Column(db.String(20), nullable=False, default='domestic')
    trip_date = db.Column(db.Date)
    requires_dl = db.Column(db.Boolean, default=True, nullable=False)
    pax = db.Column(db.Integer, nullable=False, default=1)
    created_by_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    created_by = db.relationship('User', foreign_keys=[created_by_id], lazy=True)
    members = db.relationship('KYCCustomer', backref='group_link', lazy=True,
                              foreign_keys='KYCCustomer.group_link_id')

    def completed_count(self):
        """Return number of members who have completed both KYC + indemnity.
        Phase 6c: uses derived status so stuck-flag rows count correctly."""
        return sum(
            1 for m in self.members
            if (m.deleted_at is None)
            and m.effective_kyc_submitted
            and m.effective_indemnity_signed
        )

    def kyc_done_count(self):
        """Return number of members who completed KYC form (derived)."""
        return sum(
            1 for m in self.members
            if (m.deleted_at is None) and m.effective_kyc_submitted
        )


class KYCForm(db.Model):
    __tablename__ = 'kyc_forms'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    
    # JSON field storing form fields configuration
    fields_config = db.Column(db.Text, default='[]')


class KYCSubmission(db.Model):
    __tablename__ = 'kyc_submissions'
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('kyc_customers.id'), nullable=False, unique=True)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    
    # JSON field storing submitted data
    form_data = db.Column(db.Text, default='{}')
    
    # Document paths (stored as JSON: {"passport_front": "path", "passport_back": "path", ...})
    document_paths = db.Column(db.Text, default='{}')


class IndemnityTemplate(db.Model):
    __tablename__ = 'indemnity_templates'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    
    # T&Cs content (HTML/text) - legacy field, now using PDF
    terms_content = db.Column(db.Text, default='')
    
    # T&Cs PDF file path (new)
    terms_pdf_path = db.Column(db.String(255))
    
    # Indemnity content (HTML/text) - legacy field
    indemnity_content = db.Column(db.Text, default='')
    
    # Indemnity PDF file path
    pdf_path = db.Column(db.String(255))
    
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    
    uploaded_by = db.relationship('User', foreign_keys=[uploaded_by_id], lazy=True)


class IndemnityRequest(db.Model):
    __tablename__ = 'indemnity_requests'
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('kyc_customers.id'), nullable=False)
    template_id = db.Column(db.Integer, db.ForeignKey('indemnity_templates.id'), nullable=False)
    sent_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    sent_by_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    
    # T&C acceptance tracking
    terms_accepted_at = db.Column(db.DateTime)
    terms_accepted_location = db.Column(db.String(100))
    
    # Signature data
    signed_at = db.Column(db.DateTime)
    signature_data = db.Column(db.Text)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.Text)
    
    # Phase 6c: add a backref so KYCCustomer.indemnity_requests resolves.
    # viewonly-style usage — we only read this to compute derived status.
    customer = db.relationship(
        'KYCCustomer',
        foreign_keys=[customer_id],
        backref=db.backref('indemnity_requests', lazy=True),
        lazy=True,
    )
    template = db.relationship('IndemnityTemplate', foreign_keys=[template_id], lazy=True)
    sent_by = db.relationship('User', foreign_keys=[sent_by_id], lazy=True)


def generate_token(length=32):
    """Generate a random token for secure external links."""
    import secrets
    return secrets.token_urlsafe(length)


# ==================== END KYC MODULE ====================


def ensure_schema():
    """Minimal migration to add missing columns when using SQLite."""
    try:
        eng = db.engine
        # Only run these PRAGMA-based adjustments for SQLite
        try:
            dialect = getattr(eng, 'dialect', None)
            name = getattr(dialect, 'name', '') if dialect else ''
            if name and name.lower() == 'postgresql':
                with eng.connect() as conn:
                    # Add column to tickets if missing
                    conn.execute(text("""
                        ALTER TABLE tickets
                        ADD COLUMN IF NOT EXISTS lead_assignee_id INTEGER
                    """))
                    # Create notifications table
                    conn.execute(text("""
                        CREATE TABLE IF NOT EXISTS notifications (
                            id SERIAL PRIMARY KEY,
                            user_id INTEGER NOT NULL REFERENCES users(id),
                            type VARCHAR(50),
                            ticket_id INTEGER REFERENCES tickets(id) ON DELETE SET NULL,
                            message TEXT,
                            actor_id INTEGER REFERENCES users(id),
                            created_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
                            is_read BOOLEAN NOT NULL DEFAULT FALSE
                        )
                    """))
                    # Create app_meta table
                    conn.execute(text("""
                        CREATE TABLE IF NOT EXISTS app_meta (
                            key VARCHAR(50) PRIMARY KEY,
                            value VARCHAR(200)
                        )
                    """))
                    # KYC table migrations
                    conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN IF NOT EXISTS trip_date DATE"))
                    conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN IF NOT EXISTS trip_name VARCHAR(200)"))
                    conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN IF NOT EXISTS requires_dl BOOLEAN DEFAULT TRUE"))
                    conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS deleted_at TIMESTAMP"))
                    conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN IF NOT EXISTS deleted_at TIMESTAMP"))
                    conn.execute(text("ALTER TABLE indemnity_templates ADD COLUMN IF NOT EXISTS terms_pdf_path VARCHAR(255)"))
                    conn.execute(text("ALTER TABLE indemnity_requests ADD COLUMN IF NOT EXISTS terms_accepted_at TIMESTAMP"))
                    conn.execute(text("ALTER TABLE indemnity_requests ADD COLUMN IF NOT EXISTS terms_accepted_location VARCHAR(100)"))
                    # Group KYC links table
                    conn.execute(text("""
                        CREATE TABLE IF NOT EXISTS kyc_group_links (
                            id SERIAL PRIMARY KEY,
                            token VARCHAR(64) UNIQUE NOT NULL,
                            booking_id VARCHAR(80),
                            trip_name VARCHAR(200),
                            trip_type VARCHAR(20) NOT NULL DEFAULT 'domestic',
                            trip_date DATE,
                            requires_dl BOOLEAN NOT NULL DEFAULT TRUE,
                            pax INTEGER NOT NULL DEFAULT 1,
                            created_by_id INTEGER NOT NULL REFERENCES users(id),
                            created_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW()
                        )
                    """))
                    conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN IF NOT EXISTS group_link_id INTEGER REFERENCES kyc_group_links(id)"))
                    # Phase 3: NPS feedback columns (additive, nullable).
                    conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN IF NOT EXISTS nps_rating INTEGER"))
                    conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN IF NOT EXISTS nps_feedback TEXT"))
                    conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN IF NOT EXISTS nps_submitted_at TIMESTAMP"))
                    # Helpful indexes
                    conn.execute(text("""
                        CREATE INDEX IF NOT EXISTS idx_notifications_user_created
                        ON notifications(user_id, created_at DESC)
                    """))
                    conn.execute(text("""
                        CREATE INDEX IF NOT EXISTS idx_notifications_ticket
                        ON notifications(ticket_id)
                    """))
                    conn.commit()
                return
            if name and name.lower() != 'sqlite':
                return
        except Exception:
            # If detection fails, proceed conservatively
            pass
        with eng.connect() as conn:
            cols = [row[1] for row in conn.execute(text("PRAGMA table_info(tickets)"))]
            if 'notify_emails' not in cols:
                conn.execute(text("ALTER TABLE tickets ADD COLUMN notify_emails TEXT"))
                conn.commit()
            if 'updated_at' not in cols:
                # Provide a server default to satisfy NOT NULL on existing rows
                conn.execute(text("ALTER TABLE tickets ADD COLUMN updated_at DATETIME DEFAULT (CURRENT_TIMESTAMP) NOT NULL"))
                conn.commit()
            if 'resolved_at' not in cols:
                conn.execute(text("ALTER TABLE tickets ADD COLUMN resolved_at DATETIME"))
                conn.commit()
            if 'resolved_by' not in cols:
                conn.execute(text("ALTER TABLE tickets ADD COLUMN resolved_by VARCHAR(120)"))
                conn.commit()
            if 'lead_assignee_id' not in cols:
                conn.execute(text("ALTER TABLE tickets ADD COLUMN lead_assignee_id INTEGER"))
                conn.commit()
            # Backfill any NULL updated_at values (covers legacy rows or prior schema)
            try:
                conn.execute(text("UPDATE tickets SET updated_at = CURRENT_TIMESTAMP WHERE updated_at IS NULL"))
                conn.commit()
            except Exception:
                pass
            
            # Add trip_date and trip_name columns to kyc_customers table
            kyc_cols = [row[1] for row in conn.execute(text("PRAGMA table_info(kyc_customers)"))]
            if 'trip_date' not in kyc_cols:
                conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN trip_date DATE"))
                conn.commit()
                print("[info] Added trip_date column to kyc_customers table")
            if 'trip_name' not in kyc_cols:
                conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN trip_name VARCHAR(200)"))
                conn.commit()
                print("[info] Added trip_name column to kyc_customers table")
            if 'requires_dl' not in kyc_cols:
                conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN requires_dl BOOLEAN DEFAULT 1"))
                conn.commit()
                print("[info] Added requires_dl column to kyc_customers table")
            if 'deleted_at' not in kyc_cols:
                conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN deleted_at DATETIME"))
                conn.commit()
                print("[info] Added deleted_at column to kyc_customers table")
            if 'deleted_at' not in cols:
                conn.execute(text("ALTER TABLE tickets ADD COLUMN deleted_at DATETIME"))
                conn.commit()
                print("[info] Added deleted_at column to tickets table")
            
            # Add terms_pdf_path column to indemnity_templates table
            template_cols = [row[1] for row in conn.execute(text("PRAGMA table_info(indemnity_templates)"))]
            if 'terms_pdf_path' not in template_cols:
                conn.execute(text("ALTER TABLE indemnity_templates ADD COLUMN terms_pdf_path VARCHAR(255)"))
                conn.commit()
                print("[info] Added terms_pdf_path column to indemnity_templates table")
            
            # Add T&C acceptance tracking columns to indemnity_requests table
            request_cols = [row[1] for row in conn.execute(text("PRAGMA table_info(indemnity_requests)"))]
            if 'terms_accepted_at' not in request_cols:
                conn.execute(text("ALTER TABLE indemnity_requests ADD COLUMN terms_accepted_at DATETIME"))
                conn.commit()
                print("[info] Added terms_accepted_at column to indemnity_requests table")
            if 'terms_accepted_location' not in request_cols:
                conn.execute(text("ALTER TABLE indemnity_requests ADD COLUMN terms_accepted_location VARCHAR(100)"))
                conn.commit()
                print("[info] Added terms_accepted_location column to indemnity_requests table")
            
            # Add group_link_id to kyc_customers for group KYC links
            if 'group_link_id' not in kyc_cols:
                conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN group_link_id INTEGER"))
                conn.commit()
                print("[info] Added group_link_id column to kyc_customers table")

            # Phase 3: NPS feedback columns (additive, nullable).
            if 'nps_rating' not in kyc_cols:
                conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN nps_rating INTEGER"))
                conn.commit()
                print("[info] Added nps_rating column to kyc_customers table")
            if 'nps_feedback' not in kyc_cols:
                conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN nps_feedback TEXT"))
                conn.commit()
                print("[info] Added nps_feedback column to kyc_customers table")
            if 'nps_submitted_at' not in kyc_cols:
                conn.execute(text("ALTER TABLE kyc_customers ADD COLUMN nps_submitted_at DATETIME"))
                conn.commit()
                print("[info] Added nps_submitted_at column to kyc_customers table")

        # -----------------------------------------------------------
        # Phase 6c: ONE-TIME STATUS BACKFILL (idempotent, heal-only)
        # -----------------------------------------------------------
        # Flips `kyc_submitted` / `indemnity_signed` from FALSE -> TRUE when
        # the underlying source row proves the event happened but the flag
        # drifted. This can happen due to:
        #   - transient MVCC lag during Phase 1's commit+verify,
        #   - pre-Phase-1 historical rows,
        #   - concurrent group-booking submits,
        # so the fix is to trust the source row over the boolean.
        #
        # Safety guarantees:
        #   * Only ever sets FALSE -> TRUE. Never the other way. Never
        #     deletes rows. Never overwrites a non-NULL timestamp.
        #   * Scoped by `EXISTS (source row)` → cannot heal rows whose
        #     source data is missing.
        #   * Idempotent: re-running the same SQL is a no-op because the
        #     WHERE clause already excludes TRUE rows.
        #   * Gated by an AppMeta marker so we log the result exactly once.
        #     On subsequent boots it still executes (that's fine — it's a
        #     no-op) but won't spam logs.
        try:
            _heal_marker = None
            try:
                _heal_marker = AppMeta.query.filter_by(key='kyc_status_heal_v1').first()
            except Exception:
                # AppMeta may not be queryable yet on brand-new DBs — that's
                # fine, we'll proceed and try to write the marker at end.
                _heal_marker = None

            with db.engine.begin() as conn:  # one transaction for both heals
                kyc_rows = conn.execute(text(
                    """
                    UPDATE kyc_customers
                       SET kyc_submitted = TRUE,
                           kyc_submitted_at = COALESCE(
                               kyc_submitted_at,
                               (SELECT submitted_at FROM kyc_submissions
                                 WHERE kyc_submissions.customer_id = kyc_customers.id)
                           )
                     WHERE kyc_submitted = FALSE
                       AND EXISTS (SELECT 1 FROM kyc_submissions
                                    WHERE kyc_submissions.customer_id = kyc_customers.id)
                    """
                )).rowcount

                indem_rows = conn.execute(text(
                    """
                    UPDATE kyc_customers
                       SET indemnity_signed = TRUE,
                           indemnity_signed_at = COALESCE(
                               indemnity_signed_at,
                               (SELECT MAX(signed_at) FROM indemnity_requests
                                 WHERE indemnity_requests.customer_id = kyc_customers.id
                                   AND signed_at IS NOT NULL)
                           )
                     WHERE indemnity_signed = FALSE
                       AND EXISTS (SELECT 1 FROM indemnity_requests
                                    WHERE indemnity_requests.customer_id = kyc_customers.id
                                      AND signed_at IS NOT NULL)
                    """
                )).rowcount

            if _heal_marker is None and (kyc_rows or indem_rows):
                # First-ever run that actually healed something — log + record.
                print(f"[heal] kyc_status_heal_v1: kyc_submitted healed={kyc_rows}, "
                      f"indemnity_signed healed={indem_rows}")
                try:
                    db.session.add(AppMeta(
                        key='kyc_status_heal_v1',
                        value=datetime.utcnow().isoformat()
                    ))
                    db.session.commit()
                except Exception as _e:
                    db.session.rollback()
                    print(f"[warn] failed to record kyc_status_heal_v1 marker: {_e}")
            elif _heal_marker is None:
                # Nothing to heal on first boot — still drop the marker to
                # keep future boots quiet.
                try:
                    db.session.add(AppMeta(
                        key='kyc_status_heal_v1',
                        value=datetime.utcnow().isoformat()
                    ))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
            else:
                # Marker already exists; only log if the idempotent re-run
                # still heals rows (which would mean new drift appeared).
                if kyc_rows or indem_rows:
                    print(f"[heal] kyc_status_heal_v1 (repeat): "
                          f"kyc={kyc_rows}, indem={indem_rows}")
        except Exception as _e:
            print(f"[warn] Phase 6c heal failed (non-fatal): {_e}")

    except Exception as e:
        print(f"[warn] ensure_schema failed: {e}")


def migrate_arrival_departure_fields():
    """One-time migration: split old combined arrival_time/departure_time 
    (datetime-local format like '2025-04-15T10:00') into separate 
    arrival_date + arrival_time and departure_date + departure_time keys."""
    import json
    try:
        submissions = KYCSubmission.query.all()
        migrated = 0
        for sub in submissions:
            if not sub.form_data:
                continue
            try:
                data = json.loads(sub.form_data)
            except:
                continue
            changed = False
            
            # Check if arrival_time has a combined datetime-local value (contains 'T')
            arr_time = data.get('arrival_time', '')
            if arr_time and 'T' in arr_time and not data.get('arrival_date'):
                # Split '2025-04-15T10:00' into date + time
                parts = arr_time.split('T')
                if len(parts) == 2:
                    date_str = parts[0]  # '2025-04-15'
                    time_str = parts[1]  # '10:00'
                    # Convert date from YYYY-MM-DD to DD-MM-YYYY
                    try:
                        dp = date_str.split('-')
                        data['arrival_date'] = f"{dp[2]}-{dp[1]}-{dp[0]}"
                    except:
                        data['arrival_date'] = date_str
                    # Convert time from 24hr to 12hr
                    try:
                        tp = time_str.split(':')
                        h, m = int(tp[0]), tp[1]
                        ampm = 'PM' if h >= 12 else 'AM'
                        h = h % 12 or 12
                        data['arrival_time'] = f"{h}:{m} {ampm}"
                    except:
                        data['arrival_time'] = time_str
                    changed = True
            
            dep_time = data.get('departure_time', '')
            if dep_time and 'T' in dep_time and not data.get('departure_date'):
                parts = dep_time.split('T')
                if len(parts) == 2:
                    date_str = parts[0]
                    time_str = parts[1]
                    try:
                        dp = date_str.split('-')
                        data['departure_date'] = f"{dp[2]}-{dp[1]}-{dp[0]}"
                    except:
                        data['departure_date'] = date_str
                    try:
                        tp = time_str.split(':')
                        h, m = int(tp[0]), tp[1]
                        ampm = 'PM' if h >= 12 else 'AM'
                        h = h % 12 or 12
                        data['departure_time'] = f"{h}:{m} {ampm}"
                    except:
                        data['departure_time'] = time_str
                    changed = True
            
            if changed:
                sub.form_data = json.dumps(data)
                migrated += 1
        
        if migrated:
            db.session.commit()
            print(f"[info] Migrated {migrated} submission(s): split arrival/departure datetime fields")
    except Exception as e:
        print(f"[warn] arrival/departure migration failed: {e}")


with app.app_context():
    db.create_all()
    ensure_schema()
    migrate_arrival_departure_fields()
    # One-time IST migration (optional, guarded by env var and meta flag)
    try:
        do_ist = (os.getenv('APPLY_IST_MIGRATION') or '').strip().lower() in {'1','true','yes','on'}
        done = db.session.get(AppMeta, 'ist_shift_done') is not None
        if do_ist and not done:
            shift = timedelta(hours=5, minutes=30)
            # Shift tickets
            try:
                for t in Ticket.query.all():
                    if t.created_at:
                        t.created_at = t.created_at + shift
                    if t.updated_at:
                        t.updated_at = t.updated_at + shift
                    if t.resolved_at:
                        t.resolved_at = t.resolved_at + shift
                db.session.commit()
            except Exception as e:
                print(f"[warn] IST shift tickets failed: {e}")
            # Shift comments
            try:
                for c in Comment.query.all():
                    if c.created_at:
                        c.created_at = c.created_at + shift
                db.session.commit()
            except Exception as e:
                print(f"[warn] IST shift comments failed: {e}")
            try:
                db.session.add(AppMeta(key='ist_shift_done', value=datetime.utcnow().isoformat()))
                db.session.commit()
            except Exception as e:
                print(f"[warn] IST shift meta save failed: {e}")
    except Exception as e:
        print(f"[warn] IST migration block failed: {e}")
    # Bootstrap initial admin if no users exist
    if User.query.count() == 0:
        admin_email = os.getenv('ADMIN_EMAIL', 'admin@deyor.local')
        admin_pass = os.getenv('ADMIN_PASSWORD', 'admin123')
        admin = User(name='Admin', email=admin_email, role='admin', department='International Operations')
        admin.set_password(admin_pass)
        db.session.add(admin)
        db.session.commit()

    # Optional: one-time password reset and role promotion via environment variables
    try:
        reset_email = (os.getenv('RESET_USER_EMAIL') or '').strip()
        reset_pass = (os.getenv('RESET_USER_PASSWORD') or '').strip()
        reset_make_admin = (os.getenv('RESET_USER_MAKE_ADMIN') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
        if reset_email and reset_pass:
            u = None
            try:
                u = User.query.filter(func.lower(User.email) == reset_email.lower()).first()
            except Exception as e:
                print(f"[warn] reset email lookup failed for {reset_email}: {e}")
            if u:
                u.set_password(reset_pass)
                db.session.commit()
                print(f"[info] Password reset via env for user {u.email}")
            else:
                print(f"[warn] RESET_USER_EMAIL specified but no user found: {reset_email}")

        # Allow admin promotion separately (does not require providing a password)
        if reset_email and reset_make_admin:
            u2 = None
            try:
                u2 = User.query.filter(func.lower(User.email) == reset_email.lower()).first()
            except Exception as e:
                print(f"[warn] admin promotion lookup failed for {reset_email}: {e}")
            if u2:
                if u2.role != 'admin':
                    u2.role = 'admin'
                    db.session.commit()
                    print(f"[info] Role promotion via env: {u2.email} -> admin")
                else:
                    print(f"[info] Role promotion skipped: {u2.email} already admin")
            else:
                print(f"[warn] RESET_USER_MAKE_ADMIN set but no user found for: {reset_email}")
    except Exception as e:
        print(f"[warn] env-based password reset failed: {e}")


# Login manager setup
login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ---- Jinja filter to format datetimes in IST ----
def fmt_dt_ist(dt: datetime, fmt: str = '%Y-%m-%d %H:%M') -> str:
    if not dt:
        return '—'
    try:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo('UTC'))
        return dt.astimezone(ZoneInfo('Asia/Kolkata')).strftime(fmt)
    except Exception:
        try:
            return dt.strftime(fmt)
        except Exception:
            return str(dt)

app.jinja_env.filters['fmt_ist'] = fmt_dt_ist


# ---- Navbar notifications context ----
@app.context_processor
def inject_nav_notifications():
    try:
        if current_user.is_authenticated:
            # Scope navbar dropdown to ONLY the current user's notifications to avoid duplicates
            base_q = Notification.query.filter_by(user_id=current_user.id)
            # Show only unread items in the bell dropdown
            notes = (
                base_q
                .filter_by(is_read=False)
                .order_by(Notification.created_at.desc())
                .limit(10)
                .all()
            )
            unread_count = base_q.filter_by(is_read=False).count()
            return dict(nav_notifications=notes, nav_unread_count=unread_count)
    except Exception as e:
        print(f"[warn] inject_nav_notifications failed: {e}")
    return dict(nav_notifications=[], nav_unread_count=0)


@app.route('/')
def index():
    # If served on support.deyor.in, land users directly on the public submit form
    try:
        host = (request.host or '').lower()
    except Exception:
        host = ''
    if 'support.deyor.in' in host:
        return redirect(url_for('public_submit'))
    return redirect(url_for('list_tickets'))


@app.route('/tickets')
@login_required
def list_tickets():
    team = request.args.get('team', '')
    status = request.args.get('status', '')
    q = request.args.get('q', '').strip()
    date_range = request.args.get('date_range', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()

    # Admin sees all, agents see only tickets assigned to them
    # Always exclude soft-deleted tickets from normal listing
    show_deleted = request.args.get('show_deleted', '') == '1' and getattr(current_user, 'role', None) == 'admin'
    if getattr(current_user, 'role', None) == 'admin':
        query = Ticket.query
    else:
        query = Ticket.query.join(ticket_assignees).filter(ticket_assignees.c.user_id == current_user.id)
    if not show_deleted:
        query = query.filter(Ticket.deleted_at.is_(None))
    if team:
        query = query.filter(Ticket.team == team)
    if status:
        query = query.filter(Ticket.status == status)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Ticket.subject.ilike(like),
                Ticket.description.ilike(like),
                Ticket.customer_name.ilike(like),
                Ticket.contact.ilike(like),
                Ticket.booking_id.ilike(like),
            )
        )

    # Date range filtering
    start_dt = None
    end_dt = None
    now = datetime.utcnow()
    if date_range == 'today':
        start_dt = datetime(now.year, now.month, now.day)
        end_dt = now
    elif date_range == 'yesterday':
        y = now - timedelta(days=1)
        start_dt = datetime(y.year, y.month, y.day)
        end_dt = datetime(now.year, now.month, now.day)
    elif date_range == 'last7':
        start_dt = now - timedelta(days=7)
        end_dt = now
    elif date_range == 'this_month':
        start_dt = datetime(now.year, now.month, 1)
        end_dt = now
    elif date_range == 'last_month':
        # First day of this month -> go back one day to get last month
        first_this_month = datetime(now.year, now.month, 1)
        last_month_last_day = first_this_month - timedelta(days=1)
        start_dt = datetime(last_month_last_day.year, last_month_last_day.month, 1)
        end_dt = first_this_month
    elif date_range == 'custom':
        try:
            if start_date:
                y, m, d = [int(x) for x in start_date.split('-')]
                start_dt = datetime(y, m, d)
            if end_date:
                y2, m2, d2 = [int(x) for x in end_date.split('-')]
                # end of day
                end_dt = datetime(y2, m2, d2, 23, 59, 59)
        except Exception:
            pass

    if start_dt:
        query = query.filter(Ticket.created_at >= start_dt)
    if end_dt:
        query = query.filter(Ticket.created_at <= end_dt)

    # Compute summary counts on the (possibly) filtered base query
    total_count = query.count()
    open_count = query.filter(Ticket.status == 'Open').count()
    inprog_count = query.filter(Ticket.status == 'In Progress').count()
    resolved_count = query.filter(Ticket.status == 'Resolved').count()

    # Derived metrics
    closed_pct = None
    if total_count > 0:
        try:
            closed_pct = round((resolved_count / total_count) * 100)
        except Exception:
            closed_pct = None

    # Average time to close among resolved tickets within the filtered query
    avg_close_human = '—'
    try:
        resolved_tickets = query.filter(Ticket.status == 'Resolved', Ticket.resolved_at.isnot(None)).all()
        deltas = [
            (t.resolved_at - t.created_at).total_seconds()
            for t in resolved_tickets
            if t.resolved_at and t.created_at and t.resolved_at >= t.created_at
        ]
        if deltas:
            avg_sec = sum(deltas) / len(deltas)
            minutes = int(avg_sec // 60)
            if minutes < 60:
                avg_close_human = f"{minutes}m"
            else:
                hours = minutes // 60
                minutes = minutes % 60
                if hours < 24:
                    avg_close_human = f"{hours}h {minutes}m"
                else:
                    days = hours // 24
                    hours = hours % 24
                    avg_close_human = f"{days}d {hours}h"
    except Exception:
        avg_close_human = '—'

    tickets = query.order_by(Ticket.created_at.desc()).all()
    # Notifications for the dashboard bar
    try:
        # Admins and Founders (by email) see all notifications
        founder_emails = set(e[0].lower() for e in db.session.query(Founder.email).all())
        is_founder_user = (getattr(current_user, 'email', '') or '').lower() in founder_emails
        if getattr(current_user, 'role', None) == 'admin' or is_founder_user:
            notes = Notification.query.order_by(Notification.created_at.desc()).limit(20).all()
        else:
            notes = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).limit(20).all()
        unread_count = sum(1 for n in notes if not n.is_read)
    except Exception:
        notes = []
        unread_count = 0
    return render_template(
        'tickets_list.html',
        tickets=tickets,
        show_deleted=show_deleted,
        TEAMS=TEAMS,
        PRIORITIES=PRIORITIES,
        STATUSES=STATUSES,
        current_team=team,
        current_status=status,
        q=q,
        total_count=total_count,
        open_count=open_count,
        inprog_count=inprog_count,
        resolved_count=resolved_count,
        closed_pct=closed_pct,
        avg_close_time=avg_close_human,
        current_date_range=date_range,
        start_date=start_date,
        end_date=end_date,
        notifications=notes,
        unread_count=unread_count,
    )


@app.route('/tickets/new', methods=['GET', 'POST'])
@login_required
def new_ticket():
    if request.method == 'POST':
        description = request.form.get('description', '').strip()
        team = request.form.get('team', '').strip() or 'Other'
        priority = request.form.get('priority', 'Medium')
        source = request.form.get('source', '').strip()
        customer_name = request.form.get('customer_name', '').strip()
        contact = request.form.get('contact', '').strip()
        booking_id = request.form.get('booking_id', '').strip()
        assigned_to = request.form.get('assigned_to', '').strip()  # legacy single assignee
        assignee_ids_raw = request.form.getlist('assignees')  # from multi-select/checkboxes
        assignee_ids = [int(x) for x in assignee_ids_raw if x and str(x).isdigit()]
        notify_emails = request.form.get('notify_emails', '').strip()  # legacy
        lead_assignee_raw = (request.form.get('lead_assignee') or '').strip()
        lead_assignee_id = int(lead_assignee_raw) if lead_assignee_raw.isdigit() else None

        # Lead assignee is mandatory
        if not lead_assignee_id:
            flash('Please select a lead assignee.', 'danger')
            return redirect(url_for('new_ticket'))

        # Validate: if a lead is chosen, it must be in selected assignees
        if lead_assignee_id and (lead_assignee_id not in assignee_ids):
            flash('Lead assignee must be one of the selected assignees.', 'danger')
            return redirect(url_for('new_ticket'))

        if not booking_id or not description:
            flash('Booking ID and description are required.', 'danger')
            return redirect(url_for('new_ticket'))

        t = Ticket(
            subject=booking_id,
            description=description,
            team=team if team in TEAMS else 'Domestic Operations',
            priority=priority if priority in PRIORITIES else 'Medium',
            source=source,
            customer_name=customer_name,
            contact=contact,
            booking_id=booking_id,
            assigned_to=assigned_to,
            notify_emails=notify_emails or None,
        )
        # Ensure updated_at is populated on insert to satisfy NOT NULL in SQLite
        t.updated_at = datetime.utcnow()
        if lead_assignee_id:
            t.lead_assignee_id = lead_assignee_id
        # Attach selected assignees
        users = []
        if assignee_ids:
            users = User.query.filter(User.id.in_(assignee_ids)).all()
            for u in users:
                t.assignees.append(u)
        db.session.add(t)
        db.session.commit()
        try:
            notify_created(t)
        except Exception as e:
            print(f"[warn] Failed to send creation notification: {e}")
        # Create in-app notifications
        try:
            create_notifications_for_event('ticket_created', t, f"New ticket #{t.id} created", actor=current_user if getattr(current_user, 'is_authenticated', False) else None)
        except Exception as e:
            print(f"[warn] Failed to create in-app notifications: {e}")
        # Send customer confirmation if email provided or derivable from contact
        try:
            cust_email = extract_email_from_contact(contact)
            if cust_email:
                notify_customer_created(t, cust_email)
        except Exception as e:
            print(f"[warn] Failed to send customer creation email: {e}")
        flash(f'Ticket #{t.id} created.', 'success')
        return redirect(url_for('ticket_detail', ticket_id=t.id))

    agents = User.query.order_by(User.name.asc()).all()
    return render_template('ticket_form.html', TEAMS=TEAMS, PRIORITIES=PRIORITIES, agents=agents)


@app.route('/tickets/<int:ticket_id>')
@login_required
def ticket_detail(ticket_id: int):
    t = Ticket.query.get_or_404(ticket_id)
    if getattr(current_user, 'role', None) != 'admin':
        # Ensure agent is assigned
        assigned_ids = {u.id for u in t.assignees}
        if current_user.id not in assigned_ids:
            flash('You do not have access to this ticket.', 'danger')
            return redirect(url_for('list_tickets'))
    agents = User.query.order_by(User.name.asc()).all()
    return render_template('ticket_detail.html', t=t, STATUSES=STATUSES, agents=agents)


@app.route('/tickets/<int:ticket_id>/comment', methods=['POST'])
@login_required
def add_comment(ticket_id: int):
    t = Ticket.query.get_or_404(ticket_id)
    body = request.form.get('body', '').strip()
    if not body:
        flash('Comment cannot be empty.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))
    author = getattr(current_user, 'name', None)
    c = Comment(ticket_id=t.id, author=author or None, body=body)
    db.session.add(c)
    db.session.commit()
    # In-app notifications for comment
    try:
        create_notifications_for_event('comment_added', t, f"New comment on ticket #{t.id}", actor=current_user)
    except Exception as e:
        print(f"[warn] Failed to create comment notifications: {e}")
    flash('Comment added.', 'success')
    return redirect(url_for('ticket_detail', ticket_id=ticket_id))

 


@app.route('/tickets/<int:ticket_id>/resolve', methods=['POST'])
@login_required
def resolve_ticket(ticket_id: int):
    t = Ticket.query.get_or_404(ticket_id)
    if t.status != 'Resolved':
        t.status = 'Resolved'
        t.resolved_at = datetime.utcnow()
        t.resolved_by = getattr(current_user, 'name', None)
        db.session.commit()
        try:
            notify_resolved(t)
        except Exception as e:
            # Do not block resolution on email errors
            print(f"[warn] Failed to send notification: {e}")
        # In-app notifications
        try:
            create_notifications_for_event('ticket_resolved', t, f"Ticket #{t.id} resolved", actor=current_user)
        except Exception as e:
            print(f"[warn] Failed to create resolve notifications: {e}")
        # Notify customer if we have an email
        try:
            cust_email = extract_email_from_contact(t.contact)
            if cust_email:
                notify_customer_resolved(t, cust_email)
        except Exception as e:
            print(f"[warn] Failed to send customer resolution email: {e}")
        flash('Ticket marked as resolved.', 'success')
    return redirect(url_for('ticket_detail', ticket_id=ticket_id))


@app.route('/tickets/<int:ticket_id>/status', methods=['POST'])
@login_required
def update_status(ticket_id: int):
    t = Ticket.query.get_or_404(ticket_id)
    new_status = request.form.get('status', '').strip()
    actor = getattr(current_user, 'name', None)
    if new_status not in STATUSES:
        flash('Invalid status.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    t.status = new_status
    if new_status == 'Resolved':
        t.resolved_at = datetime.utcnow()
        t.resolved_by = actor
        db.session.commit()
        try:
            notify_resolved(t)
        except Exception as e:
            print(f"[warn] Failed to send notification: {e}")
        try:
            create_notifications_for_event('ticket_resolved', t, f"Ticket #{t.id} resolved", actor=current_user)
        except Exception as e:
            print(f"[warn] Failed to create resolve notifications: {e}")
        # Notify customer if we have an email
        try:
            cust_email = extract_email_from_contact(t.contact)
            if cust_email:
                notify_customer_resolved(t, cust_email)
        except Exception as e:
            print(f"[warn] Failed to send customer resolution email: {e}")
        flash('Ticket marked as resolved.', 'success')
    else:
        # Clear resolution metadata when reopening or moving to in-progress
        t.resolved_at = None
        t.resolved_by = None
        db.session.commit()
        try:
            create_notifications_for_event('status_changed', t, f"Ticket #{t.id} status: {new_status}", actor=current_user)
        except Exception as e:
            print(f"[warn] Failed to create status notifications: {e}")
        flash(f'Status updated to {new_status}.', 'success')
    return redirect(url_for('ticket_detail', ticket_id=ticket_id))


@app.route('/tickets/<int:ticket_id>/delete', methods=['POST'])
@login_required
def delete_ticket(ticket_id: int):
    if not admin_required():
        return redirect(url_for('list_tickets'))
    t = Ticket.query.get_or_404(ticket_id)
    # Soft delete - mark as deleted instead of removing
    t.deleted_at = datetime.utcnow()
    db.session.commit()
    flash(f'Ticket #{t.id} archived. It can be restored by an admin.', 'success')
    return redirect(url_for('list_tickets'))


@app.route('/tickets/<int:ticket_id>/restore', methods=['POST'])
@login_required
def restore_ticket(ticket_id: int):
    if not admin_required():
        return redirect(url_for('list_tickets'))
    t = Ticket.query.get_or_404(ticket_id)
    t.deleted_at = None
    db.session.commit()
    flash(f'Ticket #{t.id} restored.', 'success')
    return redirect(url_for('ticket_detail', ticket_id=ticket_id))


# ---- Ticket transfer (change lead assignee) ----
@app.route('/tickets/<int:ticket_id>/transfer', methods=['POST'])
@login_required
def transfer_ticket(ticket_id: int):
    t = Ticket.query.get_or_404(ticket_id)
    # Permission: must be admin or currently assigned on ticket
    if getattr(current_user, 'role', None) != 'admin':
        assigned_ids = {u.id for u in t.assignees}
        if current_user.id not in assigned_ids:
            flash('You do not have permission to transfer this ticket.', 'danger')
            return redirect(url_for('ticket_detail', ticket_id=ticket_id))
    new_lead_raw = (request.form.get('new_lead') or '').strip()
    if not new_lead_raw.isdigit():
        flash('Invalid selection.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))
    new_lead_id = int(new_lead_raw)
    try:
        new_lead = User.query.get(new_lead_id)
        if not new_lead:
            raise ValueError('User not found')
        t.lead_assignee_id = new_lead_id
        # Ensure new lead is in assignees
        if new_lead not in t.assignees:
            t.assignees.append(new_lead)
        db.session.commit()
        try:
            create_notifications_for_event('ticket_transferred', t, f"Ticket #{t.id} lead transferred to {new_lead.name}", actor=current_user)
        except Exception as e:
            print(f"[warn] Failed to create transfer notifications: {e}")
        flash('Lead transferred.', 'success')
    except Exception as e:
        print(f"[warn] transfer failed: {e}")
        flash('Transfer failed.', 'danger')
    return redirect(url_for('ticket_detail', ticket_id=ticket_id))


# ---- Notifications helpers & endpoints ----
def recipients_for_ticket(ticket: Ticket) -> list[int]:
    ids = set()
    # All assignees
    for u in ticket.assignees:
        if u and u.id:
            ids.add(u.id)
    # Lead
    if ticket.lead_assignee_id:
        ids.add(ticket.lead_assignee_id)
    # Admins get all
    try:
        for u in User.query.filter_by(role='admin').all():
            if u and u.id:
                ids.add(u.id)
        # Founders: map by email to user accounts, include them
        founder_emails = set(e[0].lower() for e in db.session.query(Founder.email).all())
        if founder_emails:
            for u in User.query.filter(func.lower(User.email).in_(founder_emails)).all():
                if u and u.id:
                    ids.add(u.id)
    except Exception:
        pass
    return list(ids)


def create_notifications_for_event(kind: str, ticket: Ticket, message: str, actor: User | None = None) -> None:
    try:
        ids = recipients_for_ticket(ticket)
        for uid in ids:
            # Deduplicate: if an unread notification of same type/ticket already exists for this user, skip creating another
            try:
                exists = (
                    Notification.query
                    .filter_by(user_id=uid, type=kind, ticket_id=ticket.id, is_read=False)
                    .first()
                )
            except Exception:
                exists = None
            if exists is None:
                n = Notification(
                    user_id=uid,
                    type=kind,
                    ticket_id=ticket.id,
                    message=message,
                    actor_id=(actor.id if actor else None),
                )
                db.session.add(n)
        db.session.commit()
    except Exception as e:
        print(f"[warn] create_notifications_for_event failed: {e}")


@app.route('/notifications/mark_all_read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    try:
        q = Notification.query.filter_by(user_id=current_user.id)
        for n in q.filter_by(is_read=False).all():
            n.is_read = True
        db.session.commit()
        flash('Notifications marked as read.', 'success')
    except Exception as e:
        print(f"[warn] mark_all_notifications_read failed: {e}")
    # Return user to the same page/dropdown context when possible
    return redirect(request.referrer or url_for('list_tickets'))


# ---- Notifications UI ----
@app.route('/notifications')
@login_required
def notifications_list():
    """List notifications for the current user with pagination.
    The bell dropdown only shows unread; this page can show unread or all.
    """
    try:
        status = (request.args.get('status') or 'unread').strip().lower()
        page = int((request.args.get('page') or '1').strip() or '1')
        per_page = 20
    except Exception:
        status = 'unread'
        page = 1
        per_page = 20

    base_q = Notification.query.filter_by(user_id=current_user.id)
    if status == 'all':
        q = base_q
    else:
        q = base_q.filter_by(is_read=False)

    total = q.count()
    pages = (total + per_page - 1) // per_page if per_page else 1
    items = (
        q.order_by(Notification.created_at.desc())
         .offset((page - 1) * per_page)
         .limit(per_page)
         .all()
    )

    return render_template(
        'notifications.html',
        items=items,
        status=status,
        page=page,
        pages=pages,
        total=total,
    )


@app.route('/notifications/<int:nid>/read', methods=['POST'])
@login_required
def mark_notification_read(nid: int):
    """Mark a single notification as read if it belongs to the current user."""
    try:
        n = Notification.query.get_or_404(nid)
        if n.user_id == current_user.id:
            n.is_read = True
            db.session.commit()
            flash('Notification marked as read.', 'success')
    except Exception as e:
        print(f"[warn] mark_notification_read failed: {e}")
    return redirect(request.referrer or url_for('notifications_list'))


@app.route('/notifications/<int:nid>/go')
@login_required
def go_notification(nid: int):
    """Open a notification target (ticket) and mark it read for the current user."""
    try:
        n = Notification.query.get_or_404(nid)
        if n.user_id == current_user.id and not n.is_read:
            n.is_read = True
            db.session.commit()
        if n.ticket_id:
            return redirect(url_for('ticket_detail', ticket_id=n.ticket_id))
    except Exception as e:
        print(f"[warn] go_notification failed: {e}")
    return redirect(url_for('notifications_list'))


def notify_resolved(ticket: Ticket) -> None:
    # Recipients: assigned users + all admins only
    recipients = set()
    # Include legacy single assignee if present
    if ticket.assigned_to and '@' in ticket.assigned_to:
        recipients.add(ticket.assigned_to.strip())
    for u in ticket.assignees:
        if u.email:
            recipients.add(u.email)
    # Include all admins
    admin_emails = [u.email for u in User.query.filter_by(role='admin').all() if u.email]
    recipients.update(admin_emails)
    # Include per-ticket extra recipients (comma-separated)
    recipients.update(parse_emails(getattr(ticket, 'notify_emails', None)))
    # Include team default recipients from env, e.g. TEAM_EMAILS_INTERNATIONAL_OPERATIONS
    recipients.update(get_team_default_emails(ticket.team))
    # Include global notify list if provided via NOTIFY_EMAILS env
    recipients.update(parse_emails(os.getenv('NOTIFY_EMAILS', '')))
    if not recipients:
        print("[info] notify_resolved: no recipients; skipping send")
        return  # notifications disabled entirely

    smtp_host = os.getenv('SMTP_HOST')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    if not smtp_host or not smtp_user or not smtp_pass:
        print("[info] notify_resolved: SMTP config incomplete; skipping send")
        return

    subject = f"Ticket #{ticket.id} resolved: {ticket.subject}"
    link = os.getenv('BASE_URL', 'http://localhost:5000') + url_for('ticket_detail', ticket_id=ticket.id)
    body = (
        f"Ticket #{ticket.id} has been resolved.\n\n"
        f"Subject: {ticket.subject}\n"
        f"Team: {ticket.team}\n"
        f"Priority: {ticket.priority}\n"
        f"Resolved by: {ticket.resolved_by or 'Unknown'} at {ticket.resolved_at}\n\n"
        f"Link: {link}\n\n"
        f"--- Description ---\n{ticket.description}\n"
    )
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = f"Undisclosed recipients <{smtp_user}>"

    context = ssl.create_default_context(cafile=certifi.where())
    if smtp_port == 465:
        with smtplib.SMTP_SSL(smtp_host, smtp_port, context=context) as server:
            server.login(smtp_user, smtp_pass)
            server.send_message(msg, from_addr=smtp_user, to_addrs=sorted(recipients))
    else:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls(context=context)
            server.login(smtp_user, smtp_pass)
            server.send_message(msg, from_addr=smtp_user, to_addrs=sorted(recipients))


def notify_customer_created(ticket: Ticket, to_email: str) -> None:
    """Send a confirmation email to the end-customer after ticket creation."""
    smtp_host = os.getenv('SMTP_HOST')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    if not smtp_host or not smtp_user or not smtp_pass:
        print("[info] notify_customer_created: SMTP config incomplete; skipping send")
        return

    subject = f"Your Deyor ticket #{ticket.id} has been created"
    body = (
        f"Hello{(' ' + ticket.customer_name) if ticket.customer_name else ''},\n\n"
        f"Thanks for reaching out. Your ticket has been created and sent to our {ticket.team} team.\n\n"
        f"Ticket ID: {ticket.id}\n"
        f"Subject: {ticket.subject}\n"
        f"Team: {ticket.team}\n"
        f"Priority: {ticket.priority}\n"
        f"Created: {ticket.created_at}\n\n"
        f"--- Details you submitted ---\n{ticket.description}\n\n"
        f"We’ll update you as soon as possible.\n"
    )
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = to_email

    context = ssl.create_default_context(cafile=certifi.where())
    if smtp_port == 465:
        with smtplib.SMTP_SSL(smtp_host, smtp_port, context=context) as server:
            server.login(smtp_user, smtp_pass)
            server.send_message(msg, from_addr=smtp_user, to_addrs=[to_email])
    else:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls(context=context)
            server.login(smtp_user, smtp_pass)
            server.send_message(msg, from_addr=smtp_user, to_addrs=[to_email])


def notify_customer_resolved(ticket: Ticket, to_email: str) -> None:
    """Send a resolution email to the end-customer when ticket is resolved."""
    smtp_host = os.getenv('SMTP_HOST')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    if not smtp_host or not smtp_user or not smtp_pass:
        print("[info] notify_customer_resolved: SMTP config incomplete; skipping send")
        return

    subject = f"Your Deyor ticket #{ticket.id} has been resolved"
    body = (
        f"Hello{(' ' + ticket.customer_name) if ticket.customer_name else ''},\n\n"
        f"We’re writing to let you know your ticket has been marked resolved.\n\n"
        f"Ticket ID: {ticket.id}\n"
        f"Subject: {ticket.subject}\n"
        f"Team: {ticket.team}\n"
        f"Resolved by: {ticket.resolved_by or 'Team'} at {ticket.resolved_at}\n\n"
        f"If this doesn’t address your concern, just reply to this email and we’ll reopen it.\n"
    )
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = to_email

    context = ssl.create_default_context(cafile=certifi.where())
    if smtp_port == 465:
        with smtplib.SMTP_SSL(smtp_host, smtp_port, context=context) as server:
            server.login(smtp_user, smtp_pass)
            server.send_message(msg, from_addr=smtp_user, to_addrs=[to_email])
    else:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls(context=context)
            server.login(smtp_user, smtp_pass)
            server.send_message(msg, from_addr=smtp_user, to_addrs=[to_email])

def parse_emails(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [e.strip() for e in raw.split(',') if e.strip()]


def extract_email_from_contact(contact: str | None) -> str | None:
    """Best-effort extraction of a single email from a free-form contact string."""
    if not contact:
        return None
    m = re.search(r'([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})', contact)
    return m.group(1) if m else None


def get_team_default_emails(team: str) -> list[str]:
    """Read default email recipients for a team from environment.
    Team name is converted to upper snake-case, e.g. 'International Operations' -> TEAM_EMAILS_INTERNATIONAL_OPERATIONS
    """
    env_key = 'TEAM_EMAILS_' + (team or '').upper().replace(' ', '_')
    return parse_emails(os.getenv(env_key, ''))


def notify_created(ticket: Ticket) -> None:
    # Recipients: assigned users + all admins only
    recipients = set()
    # Include legacy single assignee if present
    if ticket.assigned_to and '@' in ticket.assigned_to:
        recipients.add(ticket.assigned_to.strip())
    for u in ticket.assignees:
        if u.email:
            recipients.add(u.email)
    # Include all admins
    admin_emails = [u.email for u in User.query.filter_by(role='admin').all() if u.email]
    recipients.update(admin_emails)
    # Include per-ticket extra recipients (comma-separated)
    recipients.update(parse_emails(getattr(ticket, 'notify_emails', None)))
    # Include team default recipients from env, e.g. TEAM_EMAILS_INTERNATIONAL_OPERATIONS
    recipients.update(get_team_default_emails(ticket.team))
    # Include global notify list if provided via NOTIFY_EMAILS env
    recipients.update(parse_emails(os.getenv('NOTIFY_EMAILS', '')))
    if not recipients:
        print("[info] notify_created: no recipients; skipping send")
        return

    smtp_host = os.getenv('SMTP_HOST')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    if not smtp_host or not smtp_user or not smtp_pass:
        print("[info] notify_created: SMTP config incomplete; skipping send")
        return

    subject = f"New Ticket #{ticket.id}: {ticket.subject}"
    link = os.getenv('BASE_URL', 'http://localhost:5000') + url_for('ticket_detail', ticket_id=ticket.id)
    body = (
        f"A new escalation has been created.\n\n"
        f"Subject: {ticket.subject}\n"
        f"Team: {ticket.team}\n"
        f"Priority: {ticket.priority}\n"
        f"Created: {ticket.created_at}\n\n"
        f"Link: {link}\n\n"
        f"--- Details ---\n{ticket.description}\n"
    )
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = f"Undisclosed recipients <{smtp_user}>"

    context = ssl.create_default_context(cafile=certifi.where())
    if smtp_port == 465:
        with smtplib.SMTP_SSL(smtp_host, smtp_port, context=context) as server:
            server.login(smtp_user, smtp_pass)
            server.send_message(msg, from_addr=smtp_user, to_addrs=sorted(recipients))
    else:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls(context=context)
            server.login(smtp_user, smtp_pass)
            server.send_message(msg, from_addr=smtp_user, to_addrs=sorted(recipients))


# Admin: manage users (team members)
def admin_required():
    if not current_user.is_authenticated or current_user.role != 'admin':
        flash('Admin access required.', 'danger')
        return False
    return True


@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
def admin_users():
    if not admin_required():
        return redirect(url_for('list_tickets'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        role = request.form.get('role', 'agent')
        department = request.form.get('department', '')
        password = request.form.get('password', '').strip()
        if not name or not email or not password:
            flash('Name, email and password are required.', 'danger')
            return redirect(url_for('admin_users'))
        if User.query.filter_by(email=email).first():
            flash('Email already exists.', 'danger')
            return redirect(url_for('admin_users'))
        u = User(name=name, email=email, role=role, department=department)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        flash('Team member added.', 'success')
        return redirect(url_for('admin_users'))
    users = User.query.order_by(User.role.desc(), User.name.asc()).all()
    return render_template('admin_users.html', users=users, TEAMS=TEAMS)


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id: int):
    if not admin_required():
        return redirect(url_for('list_tickets'))
    u = User.query.get_or_404(user_id)
    if u.id == current_user.id:
        flash("You can't delete your own account.", 'danger')
        return redirect(url_for('admin_users'))
    if u.role == 'admin' and User.query.filter_by(role='admin').count() <= 1:
        flash("Can't delete the last admin.", 'danger')
        return redirect(url_for('admin_users'))
    # Detach from any assigned tickets and clean association rows
    try:
        db.session.execute(text("DELETE FROM ticket_assignees WHERE user_id = :uid"), {'uid': u.id})
    except Exception as e:
        print(f"[warn] failed to clean ticket_assignees for user {u.id}: {e}")
    # Nullify lead_assignee_id on tickets where this user is the lead to avoid FK violations
    try:
        db.session.execute(text("UPDATE tickets SET lead_assignee_id = NULL WHERE lead_assignee_id = :uid"), {'uid': u.id})
    except Exception as e:
        print(f"[warn] failed to nullify lead_assignee on tickets for user {u.id}: {e}")
    # Delete notifications that belong to this user
    try:
        db.session.execute(text("DELETE FROM notifications WHERE user_id = :uid"), {'uid': u.id})
    except Exception as e:
        print(f"[warn] failed to delete notifications for user {u.id}: {e}")
    # Nullify actor_id in notifications where this user was the actor
    try:
        db.session.execute(text("UPDATE notifications SET actor_id = NULL WHERE actor_id = :uid"), {'uid': u.id})
    except Exception as e:
        print(f"[warn] failed to nullify actor_id in notifications for user {u.id}: {e}")
    db.session.delete(u)
    db.session.commit()
    flash('User deleted.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/founders', methods=['GET', 'POST'])
@login_required
def admin_founders():
    if not admin_required():
        return redirect(url_for('list_tickets'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        if not name or not email:
            flash('Name and email are required.', 'danger')
            return redirect(url_for('admin_founders'))
        f = Founder(name=name, email=email)
        db.session.add(f)
        db.session.commit()
        flash('Founder added.', 'success')
        return redirect(url_for('admin_founders'))
    founders = Founder.query.order_by(Founder.name.asc()).all()
    return render_template('admin_founders.html', founders=founders)


# Auth routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        user = None
        try:
            if email:
                user = User.query.filter(func.lower(User.email) == email.lower()).first()
        except Exception as e:
            print(f"[warn] login email lookup failed for {email}: {e}")
        ok = False
        if user:
            try:
                ok = user.check_password(password)
            except Exception as e:
                # Do not leak details to the user; log and treat as invalid
                print(f"[warn] password verification error for {email}: {e}")
                ok = False
        if ok:
            login_user(user)
            return redirect(url_for('list_tickets'))
        flash('Invalid credentials.', 'danger')
        return redirect(url_for('login'))
    return render_template('login.html')


@app.route('/logout', methods=['POST'])
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# Health and readiness endpoints for Render
@app.route('/health')
def health():
    return {'status': 'ok'}


@app.route('/ready')
def ready():
    try:
        db.session.execute(text('SELECT 1'))
        return {'status': 'ready'}
    except Exception:
        return {'status': 'degraded'}, 503


# Temporary debug endpoint to verify DB connection target and data presence
@app.route('/debug/db')
def debug_db():
    try:
        # Identify current database name
        with db.engine.connect() as conn:
            dbname = conn.execute(text('SELECT current_database()')).scalar()

        # Check counts for expected tables; mark as 'missing' if table not found
        tables = ['tickets', 'users', 'comments', 'founders', 'ticket_assignees']
        counts = {}
        for t in tables:
            try:
                counts[t] = db.session.execute(text(f'SELECT COUNT(*) FROM {t}')).scalar()
            except Exception:
                counts[t] = 'missing'

        # Ticket diagnostics
        ticket_ids = [r[0] for r in db.session.execute(text('SELECT id FROM tickets ORDER BY id')).fetchall()]
        max_id = max(ticket_ids) if ticket_ids else 0
        all_ids_set = set(range(1, max_id + 1)) if max_id else set()
        missing_ids = sorted(all_ids_set - set(ticket_ids))
        
        # Check PostgreSQL sequence to see the true max ID ever assigned
        seq_val = None
        all_dbs = []
        all_tables = []
        db_created = None
        with db.engine.connect() as conn2:
            try:
                seq_val = conn2.execute(text("SELECT last_value FROM tickets_id_seq")).scalar()
            except Exception as e:
                seq_val = f"error: {e}"
            try:
                all_dbs = [r[0] for r in conn2.execute(text("SELECT datname FROM pg_database WHERE datistemplate = false")).fetchall()]
            except Exception as e:
                all_dbs = [f"error: {e}"]
            try:
                all_tables = [r[0] for r in conn2.execute(text("SELECT tablename FROM pg_tables WHERE schemaname = 'public'")).fetchall()]
            except Exception as e:
                all_tables = [f"error: {e}"]
            try:
                db_created = str(conn2.execute(text("SELECT (pg_stat_file('base/'||oid||'/PG_VERSION')).modification FROM pg_database WHERE datname = current_database()")).scalar())
            except Exception:
                pass
        
        # KYC diagnostics
        kyc_count = 0
        kyc_sub_count = 0
        indemnity_count = 0
        try:
            kyc_count = db.session.execute(text('SELECT COUNT(*) FROM kyc_customers')).scalar()
            kyc_sub_count = db.session.execute(text('SELECT COUNT(*) FROM kyc_submissions')).scalar()
            indemnity_count = db.session.execute(text('SELECT COUNT(*) FROM indemnity_requests')).scalar()
        except:
            pass

        return {
            'database': dbname,
            'db_created': db_created,
            'counts': counts,
            'ticket_ids': ticket_ids,
            'max_ticket_id': max_id,
            'missing_ticket_ids': missing_ids,
            'sequence_last_value': seq_val,
            'all_databases': all_dbs,
            'all_tables': all_tables,
            'kyc_customers': kyc_count,
            'kyc_submissions': kyc_sub_count,
            'indemnity_requests': indemnity_count,
        }
    except Exception as e:
        return {'error': str(e)}, 500


# Temporary debug endpoint to list users (read-only)
@app.route('/debug/users')
def debug_users():
    try:
        q = (request.args.get('q') or '').strip()
        query = User.query
        if q:
            like = f"%{q.lower()}%"
            query = query.filter(func.lower(User.email).like(like) | func.lower(User.name).like(like))
        rows = query.with_entities(User.id, User.email, User.role, User.name).order_by(User.id.asc()).all()
        return {
            'users': [
                {'id': r[0], 'email': r[1], 'role': r[2], 'name': r[3]}
                for r in rows
            ]
        }
    except Exception as e:
        return {'error': str(e)}, 500

# -------- Public submission + auto-assignment helpers --------
INTERNATIONAL_KEYWORDS = {
    # Countries/regions and popular cities (non-exhaustive)
    'maldives', 'bali', 'indonesia', 'vietnam', 'thailand', 'dubai', 'uae', 'united arab emirates',
    'malaysia', 'sri lanka', 'srilanka', 'singapore', 'turkey', 'istanbul', 'antalya', 'georgia',
    'tbilisi', 'azerbaijan', 'baku', 'armenia', 'almaty', 'kazakhstan', 'mauritius', 'seychelles',
    'phuket', 'krabi', 'bangkok', 'pattaya', 'hanoi', 'da nang', 'danang', 'ho chi minh', 'saigon',
    'hoi an', 'kuala lumpur', 'langkawi', 'colombo', 'bentota', 'negombo', 'male', 'maafushi',
    'phu quoc', 'doha', 'qatar', 'oman', 'muscat', 'jordan', 'amman', 'egypt', 'cairo', 'sharm',
}

DOMESTIC_KEYWORDS = {
    # India markers
    'india', 'indian', '+91',
    # States/UTs
    'ladakh', 'jammu', 'kashmir', 'jk', 'himachal', 'uttarakhand', 'punjab', 'haryana', 'rajasthan',
    'gujarat', 'maharashtra', 'goa', 'karnataka', 'kerala', 'tamil nadu', 'andhra pradesh',
    'telangana', 'odisha', 'chhattisgarh', 'madhya pradesh', 'bihar', 'jharkhand', 'west bengal',
    'sikkim', 'assam', 'meghalaya', 'arunachal', 'nagaland', 'manipur', 'mizoram', 'tripura',
    'andaman', 'nicobar', 'puducherry', 'pondicherry', 'delhi', 'ncr',
    # Popular domestic destinations & localities
    'manali', 'shimla', 'kasol', 'spiti', 'kaza', 'kibber', 'kinnaur', 'sangla', 'kalpa', 'jibhi',
    'tirthan', 'rishikesh', 'haridwar', 'nainital', 'mussoorie', 'auli', 'kasauli', 'dharamshala',
    'mcleodganj', 'dalhousie', 'khajjiar', 'leh', 'nubra', 'pangong', 'kargil', 'sonamarg', 'gulmarg',
    'pahalgam', 'srinagar', 'goa', 'jaipur', 'udaipur', 'jodhpur', 'jaisalmer', 'agra', 'varanasi',
    'khajuraho', 'rann of kutch', 'kutch', 'somnath', 'dwarka', 'saputara', 'mumbai', 'pune',
    'lonavala', 'mahabaleshwar', 'alibaug', 'bangalore', 'bengaluru', 'mysore', 'coorg', 'ooty',
    'kodaikanal', 'pondy', 'pondicherry', 'chennai', 'hyderabad', 'ahmedabad', 'surat', 'kolkata',
    'darjeeling', 'sundarbans', 'gangtok', 'lachung', 'lachen', 'pelling', 'kaziranga', 'shillong',
    'cherrapunji', 'cherrapunjee', 'dawki', 'tawang', 'bomdila', 'ziro', 'kohima', 'imphal', 'aizawl',
    'agartala', 'bhubaneswar', 'puri', 'konark', 'gokarna', 'hampi', 'hamta', 'bir', 'billing',
    'andaman', 'port blair', 'havelock', 'neil island', 'lakshadweep', 'minicoy',
}

def _norm_text(s: str | None) -> str:
    return re.sub(r'[^a-z0-9\s]', ' ', (s or '').lower())


def detect_team_from_text(text: str) -> str:
    s = _norm_text(text)
    # Domestic if any Indian indicator matches
    for kw in DOMESTIC_KEYWORDS:
        if kw in s:
            return 'Domestic Operations'
    # International if any foreign indicator matches
    for kw in INTERNATIONAL_KEYWORDS:
        if kw in s:
            return 'International Operations'
    # Heuristics
    if 'visa' in s or 'passport' in s or 'international' in s:
        return 'International Operations'
    # Default: treat as International per brief (any other international destination)
    return 'International Operations'


def assign_all_in_team(ticket: Ticket, team: str) -> None:
    users = User.query.filter_by(department=team).all()
    for u in users:
        if u not in ticket.assignees:
            ticket.assignees.append(u)


@app.route('/submit', methods=['GET', 'POST'])
def public_submit():
    provider = (os.getenv('CAPTCHA_PROVIDER') or '').strip().lower()
    site_key = (os.getenv('CAPTCHA_SITE_KEY') or '').strip()
    secret = (os.getenv('CAPTCHA_SECRET') or '').strip()
    if request.method == 'POST':
        # Honeypot anti-spam field: real users won't see/fill this
        honeypot = (request.form.get('website') or '').strip()
        if honeypot:
            flash('Thanks! We will get back to you shortly.', 'success')
            return redirect(url_for('public_submit'))

        # Optional CAPTCHA verification (only if provider and keys are present)
        if provider in ('turnstile', 'hcaptcha') and site_key and secret:
            try:
                if provider == 'turnstile':
                    token = (request.form.get('cf-turnstile-response') or '').strip()
                    if not token:
                        raise ValueError('captcha_missing')
                    resp = requests.post(
                        'https://challenges.cloudflare.com/turnstile/v0/siteverify',
                        data={
                            'secret': secret,
                            'response': token,
                            'remoteip': request.remote_addr or ''
                        },
                        timeout=5
                    )
                    data = {}
                    try:
                        data = resp.json() if resp is not None else {}
                    except Exception:
                        data = {}
                    ok = resp.ok and data.get('success') is True
                    if not ok:
                        print(f"[warn] turnstile verify failed: status={getattr(resp, 'status_code', 'n/a')} body={data}")
                        flash('Please complete the verification and try again.', 'danger')
                        return redirect(url_for('public_submit'))
                elif provider == 'hcaptcha':
                    token = (request.form.get('h-captcha-response') or '').strip()
                    if not token:
                        raise ValueError('captcha_missing')
                    resp = requests.post(
                        'https://hcaptcha.com/siteverify',
                        data={
                            'secret': secret,
                            'response': token,
                            'remoteip': request.remote_addr or ''
                        },
                        timeout=5
                    )
                    data = {}
                    try:
                        data = resp.json() if resp is not None else {}
                    except Exception:
                        data = {}
                    ok = resp.ok and data.get('success') is True
                    if not ok:
                        print(f"[warn] hcaptcha verify failed: status={getattr(resp, 'status_code', 'n/a')} body={data}")
                        flash('Please complete the verification and try again.', 'danger')
                        return redirect(url_for('public_submit'))
            except Exception as e:
                # Fail closed with a friendly message (but do not 500)
                print(f"[warn] captcha exception: {e}")
                flash('Verification failed. Please try again.', 'danger')
                return redirect(url_for('public_submit'))

        customer_name = (request.form.get('customer_name') or '').strip()
        phone = (request.form.get('phone') or '').strip()
        email = (request.form.get('email') or '').strip()
        booking_id = (request.form.get('booking_id') or '').strip()
        destination = (request.form.get('destination') or '').strip()
        concerns = (request.form.get('concerns') or '').strip()

        if not (destination or concerns):
            flash('Please provide a destination or concerns for your escalation.', 'danger')
            return redirect(url_for('public_submit'))

        # Build a combined contact string for storage
        contact_parts = []
        if phone:
            contact_parts.append(f"Phone: {phone}")
        if email:
            contact_parts.append(f"Email: {email}")
        contact = ', '.join(contact_parts)

        subject = (booking_id or destination or (customer_name and f"Escalation from {customer_name}") or 'Escalation')[:200]
        detection_text = ' '.join([booking_id, destination, concerns, phone, email])
        team = detect_team_from_text(detection_text)

        t = Ticket(
            subject=subject,
            description=(f"Destination: {destination}\n\nConcerns: {concerns}" if destination else f"Concerns: {concerns}"),
            team=team,
            priority='Medium',
            source='Public Form',
            customer_name=customer_name or None,
            contact=contact or None,
            booking_id=booking_id or None,
        )
        t.updated_at = datetime.utcnow()
        assign_all_in_team(t, team)
        db.session.add(t)
        db.session.commit()

        try:
            notify_created(t)
        except Exception as e:
            print(f"[warn] Failed to send creation notification: {e}")
        # In-app notifications for public-created ticket
        try:
            create_notifications_for_event('ticket_created', t, f"New ticket #{t.id} created", actor=None)
        except Exception as e:
            print(f"[warn] Failed to create public ticket notifications: {e}")
        # Send customer confirmation if email provided or derivable from contact
        try:
            cust_email = (email or '').strip() or extract_email_from_contact(contact)
            if cust_email:
                notify_customer_created(t, cust_email)
        except Exception as e:
            print(f"[warn] Failed to send customer creation email: {e}")

        return render_template('public_thanks.html', ticket=t)

    return render_template('public_submit.html', CAPTCHA_PROVIDER=provider, CAPTCHA_SITE_KEY=site_key, CAPTCHA_SECRET=secret)


# ==================== KYC ROUTES ====================

# Ensure uploads directory exists
# Use UPLOAD_PATH env var for Render persistent disk, fallback to local uploads for development
UPLOAD_FOLDER = os.environ.get('UPLOAD_PATH') or os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# Ensure kyc subdirectory exists for storing files
kyc_folder = os.path.join(UPLOAD_FOLDER, 'kyc')
os.makedirs(kyc_folder, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf', 'doc', 'docx'}

def allowed_file(filename, allowed_exts=None):
    """Check if file extension is allowed."""
    if allowed_exts is None:
        allowed_exts = ALLOWED_EXTENSIONS
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_exts


def allowed_document_file(filename):
    """Check if file is a valid document (PDF or Word)."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'pdf', 'doc', 'docx'}


def allowed_image_file(filename):
    """Check if file is a valid image."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg'}


def save_uploaded_file(file, prefix='', allowed_exts=None):
    """Save uploaded file and return relative path.
    
    Args:
        file: The uploaded file object
        prefix: Prefix for the filename
        allowed_exts: Set of allowed extensions (defaults to ALLOWED_EXTENSIONS)
    """
    if allowed_exts is None:
        allowed_exts = ALLOWED_EXTENSIONS
    
    if file and allowed_file(file.filename, allowed_exts):
        from werkzeug.utils import secure_filename
        ext = secure_filename(file.filename).rsplit('.', 1)[1].lower()
        filename = f"{prefix}{generate_token(16)}.{ext}"
        # Save to kyc subfolder within UPLOAD_FOLDER
        kyc_path = os.path.join(UPLOAD_FOLDER, 'kyc')
        os.makedirs(kyc_path, exist_ok=True)
        filepath = os.path.join(kyc_path, filename)
        file.save(filepath)
        return os.path.join('uploads', 'kyc', filename)
    return None


def save_document_file(file, prefix=''):
    """Save PDF or Word document file."""
    return save_uploaded_file(file, prefix, {'pdf', 'doc', 'docx'})


@app.route('/kyc')
@login_required
def kyc_dashboard():
    """Main KYC dashboard with overview stats and alerts."""
    # Phase 6c: eager-load source relationships so the dashboard's derived-
    # status checks + group-link counters cost zero extra queries.
    from sqlalchemy.orm import joinedload, selectinload
    base = KYCCustomer.query
    if current_user.role != 'admin':
        base = base.filter_by(created_by_id=current_user.id)
    customers = (
        base.options(
            joinedload(KYCCustomer.submission),
            selectinload(KYCCustomer.indemnity_requests),
        )
        .order_by(KYCCustomer.created_at.desc())
        .all()
    )

    # Phase 6c: counters read the DERIVED status so stuck-flag customers
    # don't get miscounted as pending.
    total = len(customers)
    kyc_completed = sum(1 for c in customers if c.effective_kyc_submitted)
    indemnity_signed = sum(1 for c in customers if c.effective_indemnity_signed)
    pending_kyc = total - kyc_completed
    pending_indemnity = total - indemnity_signed

    # Get active indemnity template
    active_template = IndemnityTemplate.query.filter_by(is_active=True).first()

    # Generate alerts based on trip dates (also on derived status).
    alerts = []
    urgent_customers = []
    critical_customers = []

    for customer in customers:
        days = customer.get_days_until_trip()
        if days is not None:
            not_done = (not customer.effective_kyc_submitted) or (not customer.effective_indemnity_signed)
            if customer.is_trip_started() and not_done:
                critical_customers.append(customer)
            elif days <= 7 and days >= 0 and not_done:
                urgent_customers.append(customer)
    
    # Phase 3 / Phase 6a: NPS aggregate (read-only).
    # Scoring rubric for 1..5 scale (Phase 6a update):
    #   4 or 5    -> Promoter
    #   3         -> Passive
    #   1 or 2    -> Detractor
    # NPS = %promoters - %detractors (range: -100..+100).
    nps_rated = [c for c in customers if c.nps_rating]
    nps_total = len(nps_rated)
    nps_score = None
    nps_promoters = nps_passives = nps_detractors = 0
    nps_distribution = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    if nps_total > 0:
        for c in nps_rated:
            r = c.nps_rating
            if r in nps_distribution:
                nps_distribution[r] += 1
            if r in (4, 5):
                nps_promoters += 1
            elif r == 3:
                nps_passives += 1
            elif r in (1, 2):
                nps_detractors += 1
        nps_score = round(
            (nps_promoters / nps_total - nps_detractors / nps_total) * 100
        )

    return render_template('kyc/dashboard.html',
                         customers=customers,
                         total=total,
                         kyc_completed=kyc_completed,
                         indemnity_signed=indemnity_signed,
                         pending_kyc=pending_kyc,
                         pending_indemnity=pending_indemnity,
                         active_template=active_template,
                         urgent_customers=urgent_customers,
                         critical_customers=critical_customers,
                         nps_score=nps_score,
                         nps_total=nps_total,
                         nps_promoters=nps_promoters,
                         nps_passives=nps_passives,
                         nps_detractors=nps_detractors,
                         nps_distribution=nps_distribution)


def _apply_trip_name_filter_layered(query, raw):
    """Phase 6b: layered trip-name matching.

    Strategy:
      1. Normalize user input (strip + collapse internal whitespace).
      2. If empty → return query unchanged.
      3. Build a case-insensitive substring clause against `coalesce(trip_name, '')`
         (coalesce handles NULL trip_name rows that would otherwise be dropped).
      4. If the input has <2 tokens, the substring clause *is* the token-AND
         clause, so we just apply it and return — no extra DB round-trip.
      5. If the input has 2+ tokens, run a cheap COUNT on the substring
         match. If it returns 0 rows, fall back to a token-AND clause where
         every token must appear as a substring in trip_name (handles weird
         whitespace/invisible chars in stored values, e.g. non-breaking
         spaces, double-spaces, tabs, or hidden trailing chars that made the
         original substring miss).
      6. Otherwise, keep the substring result.

    Safety:
      * Read-only. Never mutates rows.
      * COUNT query is bounded (single scalar) and only runs for multi-token
        searches → no impact on typical single-word filters.
      * On any DB error during the COUNT, we conservatively fall back to the
        token-AND path (broader recall) rather than crash the page.
      * Returns the (possibly-filtered) query back to the caller so pagination
        and downstream filters (status, trip_type) continue to work unchanged.
    """
    normalized = ' '.join((raw or '').split())
    if not normalized:
        return query
    sub_clause = func.coalesce(KYCCustomer.trip_name, '').ilike(f"%{normalized}%")
    tokens = normalized.split()
    # Single token: substring is equivalent to token-AND; apply and return.
    if len(tokens) < 2:
        return query.filter(sub_clause)
    # Multi-token: probe substring first. Only pay for token-AND fallback
    # when substring returns zero rows.
    try:
        sub_count = (
            query.with_entities(func.count(KYCCustomer.id))
                 .filter(sub_clause)
                 .scalar()
        )
    except Exception as e:
        try:
            app.logger.warning(
                f"[trip-name] substring probe failed, falling back to token-AND: {e}"
            )
        except Exception:
            pass
        sub_count = 0
    if sub_count and sub_count > 0:
        return query.filter(sub_clause)
    # Fallback: every token must appear as a case-insensitive substring.
    tokens_clause = and_(*[
        func.coalesce(KYCCustomer.trip_name, '').ilike(f"%{t}%")
        for t in tokens
    ])
    return query.filter(tokens_clause)


@app.route('/kyc/customers')
@login_required
def kyc_customers():
    """List all KYC customers with filtering, search, pagination, and session-persisted filters."""
    filter_keys = ('status', 'trip_type', 'search', 'booking_id', 'trip_name')

    # Explicit clear — user pressed the Clear button
    if request.args.get('clear_filters'):
        session.pop('kyc_filters', None)
        return redirect(url_for('kyc_customers'))

    # Bare visits (no filter keys, no page) → restore previously saved filters
    has_filter_in_url = any(k in request.args for k in filter_keys)
    if not has_filter_in_url and 'page' not in request.args:
        stored = session.get('kyc_filters')
        if stored and any(stored.values()):
            return redirect(url_for('kyc_customers',
                                    **{k: v for k, v in stored.items() if v}))

    status = request.args.get('status', '')
    trip_type = request.args.get('trip_type', '')
    search_query = request.args.get('search', '').strip()
    booking_id_query = request.args.get('booking_id', '').strip()
    trip_name_query = request.args.get('trip_name', '').strip()

    # Persist current filter state in session whenever filter keys are in URL
    if has_filter_in_url:
        session['kyc_filters'] = {
            'status': status,
            'trip_type': trip_type,
            'search': search_query,
            'booking_id': booking_id_query,
            'trip_name': trip_name_query,
        }
    
    if current_user.role == 'admin':
        query = KYCCustomer.query
    else:
        query = KYCCustomer.query.filter_by(created_by_id=current_user.id)
    
    # Exclude soft-deleted customers
    query = query.filter(KYCCustomer.deleted_at.is_(None))
    
    # Apply search filters
    if search_query:
        query = query.filter(KYCCustomer.name.ilike(f'%{search_query}%'))
    
    if booking_id_query:
        # Support multiple booking IDs separated by comma, newline, or space
        raw_ids = re.split(r'[,\n\r\s]+', booking_id_query)
        booking_ids = [bid.strip() for bid in raw_ids if bid.strip()]
        if len(booking_ids) == 1:
            query = query.filter(KYCCustomer.booking_id.ilike(f'%{booking_ids[0]}%'))
        elif len(booking_ids) > 1:
            bid_filters = [KYCCustomer.booking_id.ilike(f'%{bid}%') for bid in booking_ids]
            query = query.filter(or_(*bid_filters))
    
    # Phase 5 + 6b: layered trip-name filter (substring → token-AND fallback).
    query = _apply_trip_name_filter_layered(query, trip_name_query)

    # Phase 6c: status filters use the derived truth. A row is "KYC done"
    # if the flag is True OR a KYCSubmission exists. "Indemnity signed" if
    # the flag is True OR a signed IndemnityRequest exists. This prevents
    # the listing from silently hiding customers whose flags drifted.
    _kyc_done_expr = or_(
        KYCCustomer.kyc_submitted.is_(True),
        db.session.query(KYCSubmission.id).filter(
            KYCSubmission.customer_id == KYCCustomer.id
        ).exists()
    )
    _indem_done_expr = or_(
        KYCCustomer.indemnity_signed.is_(True),
        db.session.query(IndemnityRequest.id).filter(
            IndemnityRequest.customer_id == KYCCustomer.id,
            IndemnityRequest.signed_at.isnot(None)
        ).exists()
    )
    if status == 'kyc_pending':
        query = query.filter(~_kyc_done_expr)
    elif status == 'kyc_completed':
        query = query.filter(_kyc_done_expr)
    elif status == 'indemnity_pending':
        query = query.filter(~_indem_done_expr)
    elif status == 'indemnity_signed':
        query = query.filter(_indem_done_expr)
    elif status == 'fully_complete':
        query = query.filter(and_(_kyc_done_expr, _indem_done_expr))

    if trip_type in ['domestic', 'international']:
        query = query.filter_by(trip_type=trip_type)
    
    # Pagination
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except (TypeError, ValueError):
        page = 1
    per_page = 10
    # Phase 6c: eager-load source rows so the template's effective_* checks
    # cost zero extra queries even on a full paginated page.
    from sqlalchemy.orm import joinedload, selectinload
    pagination = (
        query.options(
            joinedload(KYCCustomer.submission),
            selectinload(KYCCustomer.indemnity_requests),
        )
        .order_by(KYCCustomer.created_at.desc())
        .paginate(page=page, per_page=per_page, error_out=False)
    )
    customers = pagination.items
    
    return render_template('kyc/customers.html', customers=customers, 
                         pagination=pagination,
                         current_status=status, current_trip_type=trip_type,
                         search_query=search_query, booking_id_query=booking_id_query,
                         trip_name_query=trip_name_query)


@app.route('/kyc/customers/export')
@login_required
def kyc_customers_export():
    """Export customers to Excel with current filters applied."""
    import json as _json
    status = request.args.get('status', '')
    trip_type = request.args.get('trip_type', '')
    search_query = request.args.get('search', '').strip()
    booking_id_query = request.args.get('booking_id', '').strip()
    trip_name_query = request.args.get('trip_name', '').strip()

    if current_user.role == 'admin':
        query = KYCCustomer.query
    else:
        query = KYCCustomer.query.filter_by(created_by_id=current_user.id)

    # Exclude soft-deleted customers
    query = query.filter(KYCCustomer.deleted_at.is_(None))

    if search_query:
        query = query.filter(KYCCustomer.name.ilike(f'%{search_query}%'))
    if booking_id_query:
        raw_ids = re.split(r'[,\n\r\s]+', booking_id_query)
        booking_ids = [bid.strip() for bid in raw_ids if bid.strip()]
        if len(booking_ids) == 1:
            query = query.filter(KYCCustomer.booking_id.ilike(f'%{booking_ids[0]}%'))
        elif len(booking_ids) > 1:
            bid_filters = [KYCCustomer.booking_id.ilike(f'%{bid}%') for bid in booking_ids]
            query = query.filter(or_(*bid_filters))
    # Phase 5 + 6b: layered trip-name filter (substring → token-AND fallback).
    # Identical to the listing so the Excel export never diverges from screen.
    query = _apply_trip_name_filter_layered(query, trip_name_query)

    # Phase 6c: derived-status filter, identical semantics to listing.
    _kyc_done_expr = or_(
        KYCCustomer.kyc_submitted.is_(True),
        db.session.query(KYCSubmission.id).filter(
            KYCSubmission.customer_id == KYCCustomer.id
        ).exists()
    )
    _indem_done_expr = or_(
        KYCCustomer.indemnity_signed.is_(True),
        db.session.query(IndemnityRequest.id).filter(
            IndemnityRequest.customer_id == KYCCustomer.id,
            IndemnityRequest.signed_at.isnot(None)
        ).exists()
    )
    if status == 'kyc_pending':
        query = query.filter(~_kyc_done_expr)
    elif status == 'kyc_completed':
        query = query.filter(_kyc_done_expr)
    elif status == 'indemnity_pending':
        query = query.filter(~_indem_done_expr)
    elif status == 'indemnity_signed':
        query = query.filter(_indem_done_expr)
    elif status == 'fully_complete':
        query = query.filter(and_(_kyc_done_expr, _indem_done_expr))
    if trip_type in ['domestic', 'international']:
        query = query.filter_by(trip_type=trip_type)

    # Phase 6c: eager-load the source relationships so the Excel row loop
    # does not trigger N+1 queries when it reads effective_* properties.
    from sqlalchemy.orm import joinedload, selectinload
    customers = (
        query.options(
            joinedload(KYCCustomer.submission),
            selectinload(KYCCustomer.indemnity_requests),
        )
        .order_by(KYCCustomer.created_at.desc())
        .all()
    )

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KYC Customers"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    headers = [
        "Name", "Email", "Phone", "Trip Name", "Trip Type", "Trip Date",
        "Booking ID", "Pax", "Group Link", "KYC Status", "KYC Submitted At",
        "Indemnity Status", "Indemnity Signed At", "Previous Trips",
        "NPS Rating", "NPS Feedback", "NPS Submitted At", "Created At"
    ]

    # Phase 4: build email/phone → customer_id maps ONCE so the "Previous
    # Trips" column is O(N) total instead of per-row N+1 queries.
    # Read-only: a single SELECT of id/email/phone for non-deleted rows.
    from collections import defaultdict
    email_index = defaultdict(set)
    phone_index = defaultdict(set)
    try:
        all_rows = db.session.query(
            KYCCustomer.id, KYCCustomer.email, KYCCustomer.phone
        ).filter(KYCCustomer.deleted_at.is_(None)).all()
        for cid, cemail, cphone in all_rows:
            ce = (cemail or '').strip().lower()
            cp = (cphone or '').strip()
            if ce:
                email_index[ce].add(cid)
            if cp:
                phone_index[cp].add(cid)
    except Exception as e:
        print(f"[warn] previous-trips index build failed: {e}")

    def _previous_trips_for(c):
        ce = (c.email or '').strip().lower()
        cp = (c.phone or '').strip()
        matches = set()
        if ce:
            matches |= email_index.get(ce, set())
        if cp:
            matches |= phone_index.get(cp, set())
        matches.discard(c.id)
        return len(matches)
    # Also include KYC form data columns dynamically
    # Gather all unique form data keys
    all_form_keys = set()
    customer_form_data = {}
    for c in customers:
        if c.submission and c.submission.form_data:
            try:
                fd = _json.loads(c.submission.form_data)
                customer_form_data[c.id] = fd
                all_form_keys.update(fd.keys())
            except Exception:
                customer_form_data[c.id] = {}
        else:
            customer_form_data[c.id] = {}
    form_keys_sorted = sorted(all_form_keys)
    headers.extend([k.replace('_', ' ').title() for k in form_keys_sorted])

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ist = ZoneInfo('Asia/Kolkata')
    for row_idx, c in enumerate(customers, 2):
        fd = customer_form_data.get(c.id, {})
        row_data = [
            c.name,
            c.email,
            c.phone,
            c.trip_name or '',
            c.trip_type.title() if c.trip_type else '',
            c.trip_date.strftime('%d-%m-%Y') if c.trip_date else '',
            c.booking_id or '',
            c.group_link.pax if c.group_link else 1,
            url_for('kyc_group_landing', token=c.group_link.token, _external=True) if c.group_link else '',
            'Completed' if c.effective_kyc_submitted else 'Pending',
            c.kyc_submitted_at.replace(tzinfo=ZoneInfo('UTC')).astimezone(ist).strftime('%d-%m-%Y %I:%M %p') if c.kyc_submitted_at else '',
            'Signed' if c.effective_indemnity_signed else 'Pending',
            c.indemnity_signed_at.replace(tzinfo=ZoneInfo('UTC')).astimezone(ist).strftime('%d-%m-%Y %I:%M %p') if c.indemnity_signed_at else '',
            _previous_trips_for(c),
            c.nps_rating if c.nps_rating else '',
            c.nps_feedback or '',
            c.nps_submitted_at.replace(tzinfo=ZoneInfo('UTC')).astimezone(ist).strftime('%d-%m-%Y %I:%M %p') if c.nps_submitted_at else '',
            c.created_at.replace(tzinfo=ZoneInfo('UTC')).astimezone(ist).strftime('%d-%m-%Y %I:%M %p') if c.created_at else '',
        ]
        row_data.extend([fd.get(k, '') for k in form_keys_sorted])

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto-fit column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ''))
        for row_idx in range(2, len(customers) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Filename
    parts = ['KYC_Customers']
    if trip_name_query:
        parts.append(trip_name_query.replace(' ', '_'))
    if booking_id_query:
        parts.append('BID_' + booking_id_query.replace(' ', '_').replace(',', '_')[:30])
    parts.append(datetime.now(ist).strftime('%d%m%Y'))
    filename = '_'.join(parts) + '.xlsx'

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/kyc/customers/new', methods=['GET', 'POST'])
@login_required
def kyc_customer_new():
    """Add a single KYC customer."""
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        trip_type = request.form.get('trip_type', 'domestic')
        trip_name = request.form.get('trip_name', '').strip()
        trip_date_str = request.form.get('trip_date', '').strip()
        booking_id = request.form.get('booking_id', '').strip()
        requires_dl = request.form.get('requires_dl') == 'on'
        pax_str = request.form.get('pax', '1').strip()
        try:
            pax = max(1, int(pax_str)) if pax_str else 1
        except ValueError:
            pax = 1
        
        if not name or not email or not phone:
            flash('Name, email and phone are required.', 'danger')
            return redirect(url_for('kyc_customer_new'))
        
        # Parse trip date (DD-MM-YYYY format)
        trip_date = None
        if trip_date_str:
            try:
                trip_date = datetime.strptime(trip_date_str, '%d-%m-%Y').date()
            except ValueError:
                flash('Invalid trip date format. Please use DD-MM-YYYY.', 'danger')
                return redirect(url_for('kyc_customer_new'))
        
        # Create group link if pax > 1
        group_link_id = None
        if pax > 1:
            group = KYCGroupLink(
                token=generate_token(),
                booking_id=booking_id or None,
                trip_name=trip_name or None,
                trip_type=trip_type,
                trip_date=trip_date,
                requires_dl=requires_dl,
                pax=pax,
                created_by_id=current_user.id
            )
            db.session.add(group)
            db.session.flush()
            group_link_id = group.id
        
        customer = KYCCustomer(
            name=name,
            email=email,
            phone=phone,
            trip_type=trip_type,
            trip_name=trip_name or None,
            trip_date=trip_date,
            booking_id=booking_id,
            requires_dl=requires_dl,
            created_by_id=current_user.id,
            group_link_id=group_link_id,
            kyc_token=generate_token(),
            indemnity_token=generate_token()
        )
        db.session.add(customer)
        db.session.commit()
        
        # Auto-send KYC email to customer
        try:
            template = IndemnityTemplate.query.filter_by(is_active=True).first()
            if template:
                send_kyc_email(customer, template)
                flash(f'Customer {name} added and KYC email sent successfully.', 'success')
            else:
                flash(f'Customer {name} added successfully. No active template found - email not sent.', 'warning')
        except Exception as e:
            flash(f'Customer {name} added but failed to send email: {str(e)}', 'warning')
            print(f"[error] Failed to auto-send KYC email: {e}")
        
        return redirect(url_for('kyc_customers'))
    
    return render_template('kyc/customer_form.html')


@app.route('/kyc/customers/bulk', methods=['GET', 'POST'])
@login_required
def kyc_customers_bulk():
    """Bulk upload KYC customers via CSV/Excel."""
    if request.method == 'POST':
        import csv
        import io
        
        csv_file = request.files.get('csv_file')
        if not csv_file:
            flash('Please upload a CSV file.', 'danger')
            return redirect(url_for('kyc_customers_bulk'))
        
        try:
            stream = io.StringIO(csv_file.stream.read().decode('utf-8'))
            reader = csv.DictReader(stream)
            
            added = 0
            errors = []
            
            for row in reader:
                try:
                    name = row.get('name', '').strip()
                    email = row.get('email', '').strip()
                    phone = row.get('phone', '').strip()
                    trip_type = row.get('trip_type', 'domestic').strip()
                    trip_name = row.get('trip_name', '').strip()
                    trip_date_str = row.get('trip_date', '').strip()
                    booking_id = row.get('booking_id', '').strip()
                    requires_dl_str = row.get('requires_dl', 'yes').strip().lower()
                    requires_dl = requires_dl_str not in ('no', 'false', '0', 'n')
                    pax_str = row.get('pax', '1').strip()
                    try:
                        pax = max(1, int(pax_str)) if pax_str else 1
                    except ValueError:
                        pax = 1
                    
                    if not name or not email or not phone:
                        errors.append(f"Missing required fields for row: {row}")
                        continue
                    
                    # Parse trip date (DD-MM-YYYY format)
                    trip_date = None
                    if trip_date_str:
                        try:
                            trip_date = datetime.strptime(trip_date_str, '%d-%m-%Y').date()
                        except ValueError:
                            errors.append(f"Invalid trip date for {name}: {trip_date_str}. Use DD-MM-YYYY format.")
                            continue
                    
                    safe_trip_type = trip_type if trip_type in ['domestic', 'international'] else 'domestic'
                    
                    # Create group link if pax > 1
                    group_link_id = None
                    if pax > 1:
                        group = KYCGroupLink(
                            token=generate_token(),
                            booking_id=booking_id or None,
                            trip_name=trip_name or None,
                            trip_type=safe_trip_type,
                            trip_date=trip_date,
                            requires_dl=requires_dl,
                            pax=pax,
                            created_by_id=current_user.id
                        )
                        db.session.add(group)
                        db.session.flush()  # get group.id before creating customer
                        group_link_id = group.id
                    
                    customer = KYCCustomer(
                        name=name,
                        email=email,
                        phone=phone,
                        trip_type=safe_trip_type,
                        trip_name=trip_name or None,
                        trip_date=trip_date,
                        booking_id=booking_id,
                        requires_dl=requires_dl,
                        created_by_id=current_user.id,
                        group_link_id=group_link_id,
                        kyc_token=generate_token(),
                        indemnity_token=generate_token()
                    )
                    db.session.add(customer)
                    added += 1
                except Exception as e:
                    errors.append(f"Error processing row {row}: {str(e)}")
            
            db.session.commit()
            
            # Auto-send KYC emails to all newly added customers
            email_sent_count = 0
            email_errors = []
            try:
                template = IndemnityTemplate.query.filter_by(is_active=True).first()
                if template:
                    # Get all customers that were just added
                    from sqlalchemy import desc
                    recently_added = KYCCustomer.query.filter(
                        KYCCustomer.created_by_id == current_user.id
                    ).order_by(desc(KYCCustomer.created_at)).limit(added).all()
                    
                    for customer in recently_added:
                        try:
                            send_kyc_email(customer, template)
                            email_sent_count += 1
                        except Exception as e:
                            email_errors.append(f"Failed to send email to {customer.email}: {str(e)}")
                else:
                    flash(f'No active template found - emails not sent to {added} customers.', 'warning')
            except Exception as e:
                email_errors.append(f"Email sending error: {str(e)}")
            
            if added:
                if email_sent_count == added:
                    flash(f'Successfully added {added} customers and sent KYC emails to all.', 'success')
                else:
                    flash(f'Successfully added {added} customers. Emails sent to {email_sent_count} customers.', 'success')
            if errors:
                for err in errors[:5]:
                    flash(err, 'warning')
            if email_errors:
                for err in email_errors[:3]:
                    flash(err, 'warning')
            
            return redirect(url_for('kyc_customers'))
            
        except Exception as e:
            flash(f'Error processing CSV: {str(e)}', 'danger')
            return redirect(url_for('kyc_customers_bulk'))
    
    return render_template('kyc/customers_bulk.html')


@app.route('/kyc/customers/<int:customer_id>/send', methods=['POST'])
@login_required
def kyc_customer_send(customer_id):
    """Send KYC and indemnity request emails to a customer."""
    customer = KYCCustomer.query.get_or_404(customer_id)
    
    # Check permission
    if current_user.role != 'admin' and customer.created_by_id != current_user.id:
        flash('You do not have permission to send emails for this customer.', 'danger')
        return redirect(url_for('kyc_customers'))
    
    # Get active indemnity template
    template = IndemnityTemplate.query.filter_by(is_active=True).first()
    if not template:
        flash('No active indemnity template found. Please upload one first.', 'danger')
        return redirect(url_for('kyc_templates'))
    
    # Generate tokens if not present
    if not customer.kyc_token:
        customer.kyc_token = generate_token()
    if not customer.indemnity_token:
        customer.indemnity_token = generate_token()
    db.session.commit()
    
    # Send email with both links
    try:
        send_kyc_email(customer, template)
        flash(f'KYC request email sent to {customer.name} ({customer.email}).', 'success')
    except Exception as e:
        flash(f'Failed to send email: {str(e)}', 'danger')
        print(f"[error] Failed to send KYC email: {e}")
    
    return redirect(url_for('kyc_customers'))


@app.route('/kyc/customers/<int:customer_id>/delete', methods=['POST'])
@login_required
def kyc_customer_delete(customer_id):
    """Soft-delete a single KYC customer (admin only)."""
    if current_user.role != 'admin':
        flash('Only admins can delete customers.', 'danger')
        return redirect(url_for('kyc_customers'))
    
    customer = KYCCustomer.query.get_or_404(customer_id)
    customer.deleted_at = datetime.utcnow()
    db.session.commit()
    
    flash(f'Customer "{customer.name}" archived. It can be restored by an admin.', 'success')
    return redirect(url_for('kyc_customers'))


@app.route('/kyc/customers/<int:customer_id>/restore', methods=['POST'])
@login_required
def kyc_customer_restore(customer_id):
    """Restore a soft-deleted KYC customer (admin only)."""
    if current_user.role != 'admin':
        flash('Only admins can restore customers.', 'danger')
        return redirect(url_for('kyc_customers'))
    
    customer = KYCCustomer.query.get_or_404(customer_id)
    customer.deleted_at = None
    db.session.commit()
    
    flash(f'Customer "{customer.name}" restored.', 'success')
    return redirect(url_for('kyc_customer_detail', customer_id=customer_id))


@app.route('/kyc/customers/bulk-delete', methods=['POST'])
@login_required
def kyc_customers_bulk_delete():
    """Bulk soft-delete KYC customers (admin only)."""
    if current_user.role != 'admin':
        flash('Only admins can delete customers.', 'danger')
        return redirect(url_for('kyc_customers'))
    
    customer_ids = request.form.getlist('customer_ids')
    if not customer_ids:
        flash('No customers selected for deletion.', 'warning')
        return redirect(url_for('kyc_customers'))
    
    ids = [int(cid) for cid in customer_ids if cid.isdigit()]
    if not ids:
        flash('Invalid customer selection.', 'danger')
        return redirect(url_for('kyc_customers'))
    
    # Soft delete - set deleted_at timestamp
    now = datetime.utcnow()
    count = KYCCustomer.query.filter(KYCCustomer.id.in_(ids), KYCCustomer.deleted_at.is_(None)).update(
        {KYCCustomer.deleted_at: now}, synchronize_session=False
    )
    db.session.commit()
    
    flash(f'{count} customer(s) archived. They can be restored by an admin.', 'success')
    return redirect(url_for('kyc_customers'))


def get_kyc_email_settings():
    """Get KYC email settings from database or environment."""
    # Try to get from AppMeta first
    smtp_host = AppMeta.query.filter_by(key='kyc_smtp_host').first()
    smtp_port = AppMeta.query.filter_by(key='kyc_smtp_port').first()
    smtp_user = AppMeta.query.filter_by(key='kyc_smtp_user').first()
    smtp_pass = AppMeta.query.filter_by(key='kyc_smtp_pass').first()
    
    settings = {
        'smtp_host': smtp_host.value if smtp_host else os.getenv('SMTP_HOST', ''),
        'smtp_port': int(smtp_port.value if smtp_port else os.getenv('SMTP_PORT', '587')),
        'smtp_user': smtp_user.value if smtp_user else os.getenv('SMTP_USER', 'care@deyor.in'),
        'smtp_pass': smtp_pass.value if smtp_pass else os.getenv('SMTP_PASS', ''),
    }
    return settings


def send_kyc_email(customer, template):
    """Send KYC and indemnity request email to customer using configured SMTP."""
    settings = get_kyc_email_settings()
    
    kyc_url = url_for('kyc_external_form', token=customer.kyc_token, _external=True)
    indemnity_url = url_for('kyc_external_indemnity', token=customer.indemnity_token, _external=True)
    
    # Format trip date
    trip_date_str = customer.trip_date.strftime('%d-%m-%Y') if customer.trip_date else 'Not specified'
    days_until = customer.get_days_until_trip()
    days_text = f"({days_until} days from now)" if days_until is not None and days_until >= 0 else ""
    trip_name_text = f" - {customer.trip_name}" if customer.trip_name else ""
    
    subject = "Complete Your KYC and Sign Required Documents - Deyor"
    
    # Warm, personalized email with 7-day requirement
    body = f"""Dear {customer.name},

Thank you for booking your {customer.trip_type.title()} trip{trip_name_text} with Deyor! We're excited to host you on an amazing adventure.

📅 Trip Date: {trip_date_str} {days_text}

📝 IMPORTANT - MANDATORY KYC COMPLETION
To ensure a smooth travel experience, you MUST complete your KYC and sign the required documents at least 7 DAYS prior to your trip date.

Failure to complete these documents within the 7-day window may result in cancellation of your booking without refund.

Please complete the following steps:

1️⃣ Complete your KYC form:
   {kyc_url}

2️⃣ Review and sign the Terms & Conditions and Indemnity Agreement:
   {indemnity_url}

Both steps are mandatory and must be completed within 7 days before your trip.
"""

    # Add group link section if this customer is part of a group
    if customer.group_link_id:
        group = KYCGroupLink.query.get(customer.group_link_id)
        if group and group.pax > 1:
            group_url = url_for('kyc_group_landing', token=group.token, _external=True)
            body += f"""
👥 TRAVELLING WITH FRIENDS / FAMILY?
You have {group.pax} travellers in your booking. Each person MUST complete their own KYC and sign the indemnity form individually.

Please share this link with your co-travellers:
   {group_url}

They can register themselves and fill their KYC through this link.
"""

    body += """
If you have any questions or need assistance, please contact our support team:
📧 care@deyor.in
📞 +91 9870417123

We look forward to hosting you!

Best regards,
Team Deyor
https://deyor.in
"""
    
    # Check if SMTP is configured
    if not settings['smtp_host'] or not settings['smtp_user'] or not settings['smtp_pass']:
        print("[warn] KYC SMTP not configured; email would have been sent to:", customer.email)
        print("[info] Please configure SMTP settings at /kyc/settings/email")
        return False
    
    try:
        context = ssl.create_default_context(cafile=certifi.where())
        with smtplib.SMTP(settings['smtp_host'], settings['smtp_port']) as server:
            server.starttls(context=context)
            server.login(settings['smtp_user'], settings['smtp_pass'])
            
            msg = MIMEText(body)
            msg['Subject'] = subject
            msg['From'] = settings['smtp_user']
            msg['To'] = customer.email
            
            server.sendmail(settings['smtp_user'], [customer.email], msg.as_string())
            return True
    except Exception as e:
        print(f"[error] Failed to send KYC email: {e}")
        raise e


@app.route('/kyc/templates')
@login_required
def kyc_templates():
    """Manage indemnity and T&C templates."""
    templates = IndemnityTemplate.query.order_by(IndemnityTemplate.uploaded_at.desc()).all()
    return render_template('kyc/templates.html', templates=templates)


@app.route('/kyc/templates/new', methods=['GET', 'POST'])
@login_required
def kyc_template_new():
    """Upload new indemnity template with T&C and Indemnity PDFs."""
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        make_active = request.form.get('make_active') == 'on'
        
        # T&C PDF is now required
        terms_pdf = request.files.get('terms_pdf')
        if not terms_pdf or not terms_pdf.filename:
            flash('T&C document (PDF or Word) is required.', 'danger')
            return redirect(url_for('kyc_template_new'))
        
        # Indemnity PDF is required
        indemnity_pdf = request.files.get('indemnity_pdf')
        if not indemnity_pdf or not indemnity_pdf.filename:
            flash('Indemnity document (PDF or Word) is required.', 'danger')
            return redirect(url_for('kyc_template_new'))
        
        if not title:
            flash('Title is required.', 'danger')
            return redirect(url_for('kyc_template_new'))
        
        # Handle T&C PDF upload
        terms_pdf_path = save_uploaded_file(terms_pdf, 'terms_')
        if not terms_pdf_path:
            flash('Failed to upload T&C document.', 'danger')
            return redirect(url_for('kyc_template_new'))
        
        # Handle Indemnity PDF upload
        indemnity_pdf_path = save_uploaded_file(indemnity_pdf, 'indemnity_')
        if not indemnity_pdf_path:
            flash('Failed to upload Indemnity document.', 'danger')
            return redirect(url_for('kyc_template_new'))
        
        template = IndemnityTemplate(
            title=title,
            description=description,
            terms_content='',  # Legacy field
            terms_pdf_path=terms_pdf_path,
            indemnity_content='',  # Legacy field
            pdf_path=indemnity_pdf_path,
            is_active=make_active,
            uploaded_by_id=current_user.id
        )
        
        # If making this active, deactivate others
        if make_active:
            IndemnityTemplate.query.update({'is_active': False})
        
        db.session.add(template)
        db.session.commit()
        
        flash('Template uploaded successfully.', 'success')
        return redirect(url_for('kyc_templates'))
    
    return render_template('kyc/template_form.html')


@app.route('/kyc/templates/<int:template_id>/activate', methods=['POST'])
@login_required
def kyc_template_activate(template_id):
    """Activate a specific template."""
    template = IndemnityTemplate.query.get_or_404(template_id)
    
    # Deactivate all others
    IndemnityTemplate.query.update({'is_active': False})
    template.is_active = True
    db.session.commit()
    
    flash(f'Template "{template.title}" is now active.', 'success')
    return redirect(url_for('kyc_templates'))


@app.route('/kyc/templates/<int:template_id>/delete', methods=['POST'])
@login_required
def kyc_template_delete(template_id):
    """Delete a template."""
    template = IndemnityTemplate.query.get_or_404(template_id)
    
    # Check if template is being used by any indemnity requests
    used_count = IndemnityRequest.query.filter_by(template_id=template.id).count()
    if used_count > 0:
        flash(f'Cannot delete template "{template.title}" - it has been used in {used_count} customer request(s).', 'danger')
        return redirect(url_for('kyc_templates'))
    
    try:
        # Delete associated PDF files
        base_dir = os.path.dirname(os.path.abspath(__file__))
        if template.pdf_path:
            pdf_path = os.path.join(base_dir, template.pdf_path) if not template.pdf_path.startswith('/') else template.pdf_path
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
        if template.terms_pdf_path:
            terms_path = os.path.join(base_dir, template.terms_pdf_path) if not template.terms_pdf_path.startswith('/') else template.terms_pdf_path
            if os.path.exists(terms_path):
                os.remove(terms_path)
        
        title = template.title
        db.session.delete(template)
        db.session.commit()
        flash(f'Template "{title}" deleted successfully.', 'success')
    except Exception as e:
        flash(f'Failed to delete template: {str(e)}', 'danger')
    
    return redirect(url_for('kyc_templates'))


@app.route('/kyc/form/builder')
@login_required
def kyc_form_builder():
    """Build and configure the KYC form."""
    import json
    form = KYCForm.query.filter_by(is_active=True).first()
    
    if not form:
        # Create default form if none exists
        default_fields = [
            {'name': 'full_name', 'label': 'Full Name (as per ID)', 'type': 'text', 'required': True},
            {'name': 'gender', 'label': 'Gender', 'type': 'select', 'required': True,
             'options': ['Male', 'Female', 'Other', 'Prefer not to say']},
            {'name': 'dob', 'label': 'Date of Birth', 'type': 'date', 'required': True},
            {'name': 'nationality', 'label': 'Nationality', 'type': 'select', 'required': True,
             'options': ['Indian', 'American', 'British', 'Canadian', 'Australian', 'German', 'French', 'Japanese', 'Chinese', 'Singaporean', 'Malaysian', 'Thai', 'Nepalese', 'Sri Lankan', 'Bangladeshi', 'Pakistani', 'UAE', 'Other']},
            {'name': 'address_line', 'label': 'Address Line', 'type': 'text', 'required': True},
            {'name': 'city', 'label': 'City', 'type': 'text', 'required': True},
            {'name': 'state', 'label': 'State', 'type': 'select', 'required': True, 'dynamic_options': True},
            {'name': 'pincode', 'label': 'Pincode/ZIP', 'type': 'text', 'required': True},
            {'name': 'id_type', 'label': 'ID Type', 'type': 'select', 'required': True, 
             'options': ['Aadhaar Card', 'PAN Card', 'Driving License', 'Passport', 'Voter ID']},
            {'name': 'id_number', 'label': 'ID Number', 'type': 'text', 'required': True},
            {'name': 'emergency_contact_name', 'label': 'Emergency Contact Name', 'type': 'text', 'required': True},
            {'name': 'emergency_contact_phone', 'label': 'Emergency Contact Phone', 'type': 'text', 'required': True},
        ]
        
        form = KYCForm(
            name='Default KYC Form',
            is_active=True,
            fields_config=json.dumps(default_fields)
        )
        db.session.add(form)
        db.session.commit()
    
    # Parse fields for template
    fields = json.loads(form.fields_config) if form.fields_config else []
    
    return render_template('kyc/form_builder.html', form=form, fields=fields)


@app.route('/kyc/form/builder/save', methods=['POST'])
@login_required
def kyc_form_builder_save():
    """Save KYC form configuration."""
    import json
    
    form = KYCForm.query.filter_by(is_active=True).first()
    if not form:
        form = KYCForm(name='KYC Form', is_active=True)
        db.session.add(form)
    
    # Parse the form configuration from POST data
    # For simplicity, we accept a JSON string
    fields_config = request.form.get('fields_config', '[]')
    
    try:
        # Validate JSON
        config = json.loads(fields_config)
        form.fields_config = fields_config
        db.session.commit()
        flash('KYC form saved successfully.', 'success')
    except json.JSONDecodeError:
        flash('Invalid form configuration.', 'danger')
    
    return redirect(url_for('kyc_form_builder'))


# ==================== EXTERNAL CUSTOMER-FACING ROUTES ====================

@app.route('/kyc/group/<token>', methods=['GET', 'POST'])
def kyc_group_landing(token):
    """Public group landing page — friends register themselves to complete KYC."""
    import json as _json
    group = KYCGroupLink.query.filter_by(token=token).first_or_404()

    # Existing members
    members = KYCCustomer.query.filter_by(group_link_id=group.id).filter(
        KYCCustomer.deleted_at.is_(None)
    ).order_by(KYCCustomer.created_at.asc()).all()

    error_msg = None
    form_values = {'name': '', 'email': '', 'phone': ''}
    is_full = len(members) >= group.pax

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        form_values = {'name': name, 'email': email, 'phone': phone}

        # Allow existing members to re-access KYC even when group is full
        if phone:
            existing = KYCCustomer.query.filter_by(
                group_link_id=group.id, phone=phone
            ).filter(KYCCustomer.deleted_at.is_(None)).first()
            if existing:
                return redirect(url_for('kyc_external_form', token=existing.kyc_token))

        # Capacity check — block new registrations once pax is reached
        if is_full:
            error_msg = f'This group is full. All {group.pax} spots have been registered. Please contact the booker or Deyor support if you believe this is a mistake.'
        # Validation: all fields required
        elif not name or not phone or not email:
            error_msg = 'Name, phone number and email are all required.'
        elif not re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email):
            error_msg = 'Please enter a valid email address.'
        else:
            # Create new KYCCustomer linked to this group
            customer = KYCCustomer(
                name=name,
                email=email,
                phone=phone,
                trip_type=group.trip_type,
                trip_name=group.trip_name,
                trip_date=group.trip_date,
                booking_id=group.booking_id,
                requires_dl=group.requires_dl,
                created_by_id=group.created_by_id,
                group_link_id=group.id,
                kyc_token=generate_token(),
                indemnity_token=generate_token()
            )
            db.session.add(customer)
            db.session.commit()

            # Redirect to their individual KYC form
            return redirect(url_for('kyc_external_form', token=customer.kyc_token))

    return render_template('kyc/group_landing.html',
                         group=group, members=members, error_msg=error_msg,
                         form_values=form_values, is_full=is_full, hide_nav=True)


@app.route('/kyc/external/<token>', methods=['GET', 'POST'])
def kyc_external_form(token):
    """External KYC form for customers to fill."""
    customer = KYCCustomer.query.filter_by(kyc_token=token).first_or_404()
    
    # Phase 6c: honour derived status so a customer whose KYC row exists
    # but whose flag drifted never gets sent through the submission path
    # a second time (which would hit the UNIQUE constraint on KYCSubmission).
    if customer.effective_kyc_submitted:
        return render_template('kyc/external_already_complete.html',
                             message='Your KYC form has already been submitted.',
                             customer=customer,
                             hide_nav=True)
    
    form = KYCForm.query.filter_by(is_active=True).first()
    if not form:
        return render_template('kyc/external_error.html', 
                             message='KYC form is not available. Please contact support.',
                             hide_nav=True)
    
    import json
    fields = json.loads(form.fields_config) if form.fields_config else []
    
    # Add passport fields for international trips
    if customer.trip_type == 'international':
        passport_fields = [
            {'name': 'passport_number', 'label': 'Passport Number', 'type': 'text', 'required': True},
            {'name': 'passport_issue_date', 'label': 'Passport Issue Date', 'type': 'date', 'required': True},
            {'name': 'passport_expiry_date', 'label': 'Passport Expiry Date', 'type': 'date', 'required': True},
            {'name': 'passport_issuing_authority', 'label': 'Issuing Authority', 'type': 'text', 'required': True},
        ]
        fields.extend(passport_fields)
    
    if request.method == 'POST':
        # Collect form data for new fields
        form_data = {
            'full_name': request.form.get('full_name', '').strip(),
            'gender': request.form.get('gender', '').strip(),
            'age': request.form.get('age', '').strip(),
            'blood_group': request.form.get('blood_group', '').strip(),
            'phone': request.form.get('phone', '').strip(),
            'email': request.form.get('email', '').strip(),
            'nationality': request.form.get('nationality', '').strip(),
            'address_line': request.form.get('address_line', '').strip(),
            'city': request.form.get('city', '').strip(),
            'state': request.form.get('state', '').strip(),
            'pincode': request.form.get('pincode', '').strip(),
            'company_name': request.form.get('company_name', '').strip(),
            'designation': request.form.get('designation', '').strip(),
            'emergency_name': request.form.get('emergency_name', '').strip(),
            'emergency_phone': request.form.get('emergency_phone', '').strip(),
            'emergency_relation': request.form.get('emergency_relation', '').strip(),
            'hear_about': request.form.get('hear_about', '').strip(),
            'arrival_mode': request.form.get('arrival_mode', '').strip(),
            'arrival_date': request.form.get('arrival_date', '').strip(),
            'arrival_time': request.form.get('arrival_time', '').strip(),
            'departure_mode': request.form.get('departure_mode', '').strip(),
            'departure_date': request.form.get('departure_date', '').strip(),
            'departure_time': request.form.get('departure_time', '').strip(),
            # Document identification numbers
            'aadhaar_number': request.form.get('aadhaar_number', '').strip(),
            # Normalise: strip ALL whitespace (incl. internal/non-breaking) and uppercase,
            # so values land in the DB clean even though we no longer regex-validate format.
            'pan_number': re.sub(r'\s+', '', request.form.get('pan_number', '')).upper(),
            'dl_number': request.form.get('dl_number', '').strip().upper(),
            'passport_number': request.form.get('passport_number', '').strip().upper(),
        }

        # Server-side validation for document numbers
        import re as _re
        doc_errors = []
        is_intl = (customer.trip_type == 'international')

        # Aadhaar: always required, 12 digits
        if not form_data['aadhaar_number']:
            doc_errors.append('Aadhaar number is required.')
        elif not _re.match(r'^\d{12}$', form_data['aadhaar_number']):
            doc_errors.append('Aadhaar number must be exactly 12 digits.')

        # PAN: required for international, optional for domestic.
        # Format validation removed (Phase 6e): the previous regex combined with
        # CSS-only uppercase styling silently rejected valid lowercase-typed PANs
        # via the browser's pattern attribute. Field is presence-checked when
        # required and normalised (whitespace stripped, uppercased) on input.
        if is_intl and not form_data['pan_number']:
            doc_errors.append('PAN number is required for international trips.')

        # DL: required only if requires_dl
        if customer.requires_dl:
            if not form_data['dl_number']:
                doc_errors.append('Driving License number is required.')

        # Passport: required for international.
        # Format validation removed (Phase 6d): the previous Indian-only regex
        # rejected legitimate foreign passports, OCI/PIO travelers, and modern
        # alphanumeric Indian variants. We now accept any non-empty value
        # within the input maxlength. The presence check below is sufficient
        # to ensure the field is filled; downstream ops verify the document.
        if is_intl:
            if not form_data['passport_number']:
                doc_errors.append('Passport number is required for international trips.')

        if doc_errors:
            for err in doc_errors:
                flash(err, 'danger')
            return redirect(url_for('kyc_external_form', token=customer.kyc_token))
        
        # Server-side normalization: handle native mobile date (YYYY-MM-DD) and time (HH:MM) formats
        import re
        for date_key in ('arrival_date', 'departure_date'):
            val = form_data.get(date_key, '')
            if re.match(r'^\d{4}-\d{2}-\d{2}$', val):
                parts = val.split('-')
                form_data[date_key] = f"{parts[2]}-{parts[1]}-{parts[0]}"
        for time_key in ('arrival_time', 'departure_time'):
            val = form_data.get(time_key, '')
            corresponding_date_key = time_key.replace('_time', '_date')
            if re.match(r'^\d{2}:\d{2}$', val):
                h, m = int(val.split(':')[0]), val.split(':')[1]
                ampm = 'PM' if h >= 12 else 'AM'
                h = h % 12 or 12
                form_data[time_key] = f"{h}:{m} {ampm}"
            # Also handle datetime-local (YYYY-MM-DDTHH:MM) as combined value
            elif 'T' in val:
                dt_parts = val.split('T')
                if len(dt_parts) == 2:
                    d_str, t_str = dt_parts
                    dp = d_str.split('-')
                    form_data[corresponding_date_key] = f"{dp[2]}-{dp[1]}-{dp[0]}" if len(dp) == 3 else d_str
                    tp = t_str.split(':')
                    if len(tp) >= 2:
                        th, tm = int(tp[0]), tp[1]
                        tampm = 'PM' if th >= 12 else 'AM'
                        th = th % 12 or 12
                        form_data[time_key] = f"{th}:{tm} {tampm}"
        
        # Handle file uploads based on trip type
        document_paths = {}
        
        if customer.trip_type == 'international':
            # International: Passport + PAN
            passport_front = request.files.get('passport_front')
            passport_back = request.files.get('passport_back')
            pan_front = request.files.get('pan_front')
            pan_back = request.files.get('pan_back')
            
            if passport_front:
                path = save_uploaded_file(passport_front, f'passport_front_{customer.id}_')
                if path:
                    document_paths['passport_front'] = path
            if passport_back:
                path = save_uploaded_file(passport_back, f'passport_back_{customer.id}_')
                if path:
                    document_paths['passport_back'] = path
            if pan_front:
                path = save_uploaded_file(pan_front, f'pan_front_{customer.id}_')
                if path:
                    document_paths['pan_front'] = path
            if pan_back:
                path = save_uploaded_file(pan_back, f'pan_back_{customer.id}_')
                if path:
                    document_paths['pan_back'] = path
        else:
            # Domestic: Aadhaar/PAN + DL
            id_front = request.files.get('id_proof_front')
            id_back = request.files.get('id_proof_back')
            dl_front = request.files.get('dl_front')
            dl_back = request.files.get('dl_back')
            
            if id_front:
                path = save_uploaded_file(id_front, f'id_front_{customer.id}_')
                if path:
                    document_paths['id_proof_front'] = path
            if id_back:
                path = save_uploaded_file(id_back, f'id_back_{customer.id}_')
                if path:
                    document_paths['id_proof_back'] = path
            if dl_front:
                path = save_uploaded_file(dl_front, f'dl_front_{customer.id}_')
                if path:
                    document_paths['dl_front'] = path
            if dl_back:
                path = save_uploaded_file(dl_back, f'dl_back_{customer.id}_')
                if path:
                    document_paths['dl_back'] = path
        
        # Create submission
        submission = KYCSubmission(
            customer_id=customer.id,
            form_data=json.dumps(form_data),
            document_paths=json.dumps(document_paths)
        )
        db.session.add(submission)
        
        # Update customer status
        customer.kyc_submitted = True
        customer.kyc_submitted_at = datetime.utcnow()

        # Commit AND verify the flag actually persisted to the database
        # before we show the customer a success message or send any
        # completion notifications. This closes the intermittent bug
        # where the email was sent but the dashboard still showed pending.
        ok, err = _commit_and_verify_kyc_status(customer, 'kyc_submitted')
        if not ok:
            # Do NOT send success emails / notifications on an unverified write.
            # The submission row may already be in the DB; that's fine — we
            # never delete it. The customer simply gets a retry prompt.
            flash(err or 'We could not save your KYC submission. Please try again.', 'danger')
            return redirect(url_for('kyc_external_form', token=customer.kyc_token))

        # Only now — after verified persistence — emit notifications.
        try:
            create_kyc_notifications(customer, 'kyc_completed')
        except Exception as e:
            print(f"[warn] Failed to create KYC completion notifications: {e}")

        # Redirect to T&Cs/Indemnity signing page
        flash('Your KYC has been submitted successfully! Please proceed to review and sign the Terms & Conditions and Indemnity Agreement.', 'success')
        return redirect(url_for('kyc_external_indemnity', token=customer.indemnity_token))
    
    return render_template('kyc/external_form.html', 
                         customer=customer, 
                         fields=fields, 
                         is_international=customer.trip_type == 'international',
                         hide_nav=True)


@app.route('/kyc/sign/<token>', methods=['GET', 'POST'])
def kyc_external_indemnity(token):
    """External indemnity signing page."""
    customer = KYCCustomer.query.filter_by(indemnity_token=token).first_or_404()
    
    # Phase 6c: use derived status so a stuck-flag but source-signed row
    # doesn't re-enter the sign flow (which could create duplicate
    # IndemnityRequest rows / regenerate PDFs).
    if customer.effective_indemnity_signed:
        # Phase 3: if they haven't submitted NPS yet, nudge them there so
        # the feedback loop isn't lost on revisits. This never overwrites
        # anything — the feedback route is idempotent and gated on
        # nps_submitted_at.
        if customer.nps_submitted_at is None:
            return redirect(url_for('kyc_feedback', token=customer.indemnity_token))
        return render_template('kyc/external_already_complete.html',
                             message='You have already signed the Terms & Conditions and Indemnity Agreement.',
                             customer=customer,
                             hide_nav=True)
    
    template = IndemnityTemplate.query.filter_by(is_active=True).first()
    if not template:
        return render_template('kyc/external_error.html',
                             message='Required documents are not available. Please contact support.',
                             hide_nav=True)
    
    if request.method == 'POST':
        # Record signature
        accept_terms = request.form.get('accept_terms') == 'on'
        accept_indemnity = request.form.get('accept_indemnity') == 'on'
        
        if not accept_terms or not accept_indemnity:
            flash('You must accept both the Terms & Conditions and Indemnity Agreement to proceed.', 'danger')
            return redirect(url_for('kyc_external_indemnity', token=token))
        
        # Capture T&C acceptance metadata
        terms_accepted_at = request.form.get('terms_accepted_at', '').strip()
        terms_accepted_location = request.form.get('terms_accepted_location', '').strip()
        
        # Capture signature metadata
        signature_timestamp = request.form.get('signature_timestamp', '').strip()
        signature_location = request.form.get('signature_location', '').strip()
        signature_ip = request.form.get('signature_ip', '').strip() or request.remote_addr
        signature_type = request.form.get('signature_type', '').strip() or 'draw'
        signature_image = request.form.get('signature_image', '').strip()
        
        # Build signature data string with image info
        signature_data = f"e-signed-{customer.indemnity_token}|type:{signature_type}|ts:{signature_timestamp}|loc:{signature_location}|ip:{signature_ip}"
        
        # Save signature image if provided
        signature_image_path = None
        if signature_image and signature_type == 'draw' and signature_image.startswith('data:image'):
            try:
                # Save base64 image
                import base64
                from werkzeug.utils import secure_filename
                
                # Parse base64 data
                header, encoded = signature_image.split(",", 1)
                image_data = base64.b64decode(encoded)
                
                # Create filename
                sig_filename = f"sig_{secure_filename(customer.name.replace(' ', '_'))}_{generate_token(8)}.png"
                sig_dir = os.path.join(UPLOAD_FOLDER, 'kyc', 'signatures')
                os.makedirs(sig_dir, exist_ok=True)
                sig_path = os.path.join(sig_dir, sig_filename)
                
                # Save file
                with open(sig_path, 'wb') as f:
                    f.write(image_data)
                
                signature_image_path = os.path.join('uploads', 'kyc', 'signatures', sig_filename)
            except Exception as e:
                print(f"[warn] Failed to save signature image: {e}")
        elif signature_type == 'type' and signature_image:
            # For typed signatures, store the text
            signature_image_path = f"typed:{signature_image}"
        
        # Create indemnity request record with T&C tracking
        indemnity_req = IndemnityRequest(
            customer_id=customer.id,
            template_id=template.id,
            sent_by_id=customer.created_by_id,
            terms_accepted_at=datetime.fromisoformat(terms_accepted_at.replace('Z', '+00:00')) if terms_accepted_at else datetime.utcnow(),
            terms_accepted_location=terms_accepted_location,
            signed_at=datetime.utcnow(),
            signature_data=signature_data,
            ip_address=signature_ip,
            user_agent=request.headers.get('User-Agent', '')
        )
        db.session.add(indemnity_req)
        db.session.commit()  # Commit to get the ID
        
        # Save signature path to signature_data field or create separate storage
        if signature_image_path:
            # Store the path in a separate column if we had one, or append to signature_data
            # For now, we'll update the record with the image path
            indemnity_req.signature_data = f"{signature_data}|img:{signature_image_path}"
        
        db.session.commit()
        
        # Embed signature in the PDF
        signed_pdf_path = None
        if signature_image_path:
            try:
                signed_pdf_path = embed_signature_in_pdf(customer, signature_image_path, indemnity_req)
                if signed_pdf_path:
                    # Store signed PDF path in signature_data for reference
                    indemnity_req.signature_data = f"{indemnity_req.signature_data}|pdf:{signed_pdf_path}"
                    db.session.commit()
                    print(f"[info] Signature embedded in PDF: {signed_pdf_path}")
            except Exception as e:
                print(f"[warn] Failed to embed signature in PDF: {e}")
        
        # Update customer status
        customer.indemnity_signed = True
        customer.indemnity_signed_at = datetime.utcnow()

        # Commit AND verify the flag actually persisted before declaring
        # success. The IndemnityRequest row and signed PDF are preserved
        # regardless — we only guard the user-visible success path and
        # admin notifications on verified persistence.
        ok, err = _commit_and_verify_kyc_status(customer, 'indemnity_signed')
        if not ok:
            flash(err or 'We could not save your signature. Please try again.', 'danger')
            return redirect(url_for('kyc_external_indemnity', token=customer.indemnity_token))

        # Create notifications for Indemnity signing (only after verified commit)
        try:
            create_kyc_notifications(customer, 'indemnity_signed')
        except Exception as e:
            print(f"[warn] Failed to create indemnity signing notifications: {e}")

        # Phase 3: send the customer to the NPS feedback page. We use a
        # GET redirect (POST-then-GET) so that refresh/back don't resubmit
        # the signature and don't retrigger notifications.
        return redirect(url_for('kyc_feedback', token=customer.indemnity_token))
    
    # For GET request, process indemnity PDF with placeholder replacement
    processed_pdf_path = None
    if template.pdf_path:
        processed_pdf_path = process_indemnity_pdf(template.pdf_path, customer)
    
    return render_template('kyc/external_indemnity.html',
                         customer=customer,
                         template=template,
                         processed_pdf_path=processed_pdf_path,
                         hide_nav=True)


@app.route('/kyc/feedback/<token>', methods=['GET', 'POST'])
def kyc_feedback(token):
    """Phase 3: NPS feedback page. Uses indemnity_token as the lookup key
    because it's the token already embedded in the indemnity flow email.
    Idempotent: once submitted, revisits show a thank-you state and the
    POST handler is a no-op (never overwrites an existing rating)."""
    customer = KYCCustomer.query.filter_by(indemnity_token=token).first_or_404()
    already = customer.nps_submitted_at is not None

    if request.method == 'POST' and not already:
        rating_raw = (request.form.get('rating') or '').strip()
        feedback = (request.form.get('feedback') or '').strip()
        try:
            rating = int(rating_raw)
        except (TypeError, ValueError):
            rating = 0
        if rating < 1 or rating > 5:
            flash('Please select a rating between 1 and 5.', 'danger')
            return redirect(url_for('kyc_feedback', token=token))

        # Cap feedback length defensively — text column can hold much more
        # but we don't want abuse. Empty feedback is allowed.
        customer.nps_rating = rating
        customer.nps_feedback = feedback[:2000] if feedback else None
        customer.nps_submitted_at = datetime.utcnow()
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            try:
                app.logger.error(f"[nps] commit failed for customer {customer.id}: {e}")
            except Exception:
                print(f"[nps] commit failed for customer {customer.id}: {e}")
            flash('We could not save your feedback. Please try again.', 'danger')
            return redirect(url_for('kyc_feedback', token=token))
        # PRG: redirect to same page so refresh won't re-POST.
        return redirect(url_for('kyc_feedback', token=token, submitted=1))

    just_submitted = already or request.args.get('submitted') == '1'
    return render_template('kyc/external_feedback.html',
                           customer=customer,
                           already_submitted=already,
                           just_submitted=just_submitted,
                           hide_nav=True)


def process_indemnity_pdf(pdf_path, customer):
    """Process indemnity PDF by replacing placeholders with customer data.
    
    Supported placeholders:
    {{customer_name}} - Customer's full name
    {{booking_id}} - Booking ID
    {{trip_date}} - Trip date (DD-MM-YYYY format)
    {{email}} - Customer email
    {{phone}} - Customer phone
    {{trip_type}} - Domestic/International
    {{signature}} - Reserved area for signature (will be filled when signed)
    
    Returns the path to the processed PDF
    """
    import os
    from PyPDF2 import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from io import BytesIO
    from werkzeug.utils import secure_filename
    
    # Prepare placeholder values
    placeholders = {
        '{{customer_name}}': customer.name or '',
        '{{booking_id}}': customer.booking_id or 'N/A',
        '{{trip_date}}': customer.trip_date.strftime('%d-%m-%Y') if customer.trip_date else 'N/A',
        '{{email}}': customer.email or '',
        '{{phone}}': customer.phone or '',
        '{{trip_type}}': customer.trip_type.title() if customer.trip_type else 'N/A',
    }
    
    try:
        # Resolve the source PDF path
        # pdf_path is like 'uploads/kyc/indemnity_xxx.pdf'
        # Strip 'uploads/' prefix to get the relative path within UPLOAD_FOLDER
        import sys
        relative_path = pdf_path.replace('uploads/', '', 1) if pdf_path.startswith('uploads/') else pdf_path
        
        # Try UPLOAD_FOLDER first (persistent disk on Render)
        full_pdf_path = os.path.join(UPLOAD_FOLDER, relative_path)
        print(f"[process_pdf] Trying UPLOAD_FOLDER path: {full_pdf_path}", file=sys.stderr)
        
        if not os.path.exists(full_pdf_path):
            # Fallback: try base_dir/uploads/ path
            base_dir = os.path.dirname(os.path.abspath(__file__))
            full_pdf_path = os.path.join(base_dir, pdf_path)
            print(f"[process_pdf] Trying base_dir path: {full_pdf_path}", file=sys.stderr)
        
        if not os.path.exists(full_pdf_path):
            print(f"[warn] PDF not found at any location for: {pdf_path}", file=sys.stderr)
            return pdf_path
        
        # Read the PDF
        reader = PdfReader(full_pdf_path)
        writer = PdfWriter()
        
        # Get page size from first page
        first_page = reader.pages[0]
        page_width = float(first_page.mediabox.width)
        page_height = float(first_page.mediabox.height)
        
        # Process each page
        for page_num, page in enumerate(reader.pages):
            # Extract text to find placeholder positions
            text_content = page.extract_text() or ''
            
            # Create overlay canvas for this page
            packet = BytesIO()
            c = canvas.Canvas(packet, pagesize=(page_width, page_height))
            
            # Search for placeholders and overlay text
            # Note: This is a simplified approach - we overlay text at estimated positions
            # For production, you would need to parse the PDF structure more carefully
            
            found_placeholders = []
            for placeholder, value in placeholders.items():
                if placeholder in text_content:
                    found_placeholders.append((placeholder, value))
            
            if found_placeholders:
                # Set font for overlay
                c.setFont("Helvetica", 10)
                c.setFillColorRGB(0, 0, 0)  # Black text
                
                # For each found placeholder, we would ideally get exact coordinates
                # For now, we'll add a header with customer info on first page
                if page_num == 0:
                    # Add customer info header overlay
                    y_pos = page_height - 50
                    c.setFont("Helvetica-Bold", 11)
                    c.drawString(50, y_pos, f"Customer: {placeholders['{{customer_name}}']}")
                    y_pos -= 15
                    c.setFont("Helvetica", 10)
                    c.drawString(50, y_pos, f"Booking ID: {placeholders['{{booking_id}}']}")
                    y_pos -= 15
                    c.drawString(50, y_pos, f"Trip Date: {placeholders['{{trip_date}}']}")
                    y_pos -= 15
                    c.drawString(50, y_pos, f"Phone: {placeholders['{{phone}}']}")
            
            c.save()
            packet.seek(0)
            
            # Create overlay PDF
            from PyPDF2 import PdfReader as OverlayReader
            try:
                overlay = OverlayReader(packet)
                if len(overlay.pages) > 0:
                    page.merge_page(overlay.pages[0])
            except Exception as e:
                print(f"[warn] Could not merge overlay for page {page_num}: {e}")
            
            writer.add_page(page)
        
        # Create processed filename
        processed_filename = f"processed_{secure_filename(customer.name.replace(' ', '_'))}_{generate_token(8)}.pdf"
        processed_dir = os.path.join(UPLOAD_FOLDER, 'kyc', 'processed')
        os.makedirs(processed_dir, exist_ok=True)
        processed_path = os.path.join(processed_dir, processed_filename)
        
        # Write the output PDF
        with open(processed_path, 'wb') as output_file:
            writer.write(output_file)
        
        # Return relative path matching the /uploads/ route
        return os.path.join('uploads', 'kyc', 'processed', processed_filename)
        
    except Exception as e:
        print(f"[warn] PDF processing failed: {e}")
        return pdf_path


def find_pdf_fields(pdf_path):
    """Dynamically find positions of form fields in the PDF using pdfplumber.
    
    Returns coordinates in ReportLab format (origin at bottom-left).
    pdfplumber uses top-left origin, so we convert: reportlab_y = page_height - pdfplumber_y
    """
    import sys
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            last_page = pdf.pages[-1]
            page_height = float(last_page.height)
            words = last_page.extract_words()
            field_positions = {}
            
            print(f"[find_fields] Page height: {page_height}", file=sys.stderr)
            print(f"[find_fields] Found {len(words)} words on last page", file=sys.stderr)
            
            # Log all words for debugging
            for i, word in enumerate(words):
                text = word['text'].lower()
                if any(kw in text for kw in ['signature', 'full', 'name', 'date', 'place', 'execution']):
                    print(f"[find_fields] Word {i}: '{word['text']}' x0={word['x0']:.1f} top={word['top']:.1f} bottom={word['bottom']:.1f}", file=sys.stderr)
            
            # Look for "Signature" label and the underline next to it
            for i, word in enumerate(words):
                text = word['text']
                
                # Find "Signature" text position
                if text.lower().replace(':', '').strip() == 'signature' or text.lower().startswith('signature'):
                    sig_word = word
                    sig_label_right = word['x1']
                    sig_y_plumber = word['top']
                    sig_bottom_plumber = word['bottom']
                    
                    # Look for underline characters (___) on the same line after "Signature"
                    underline_x0 = sig_label_right + 5
                    underline_x1 = sig_label_right + 180  # default width
                    found_underline = False
                    
                    # Method 1: Look for underscore text characters
                    for j in range(i + 1, min(i + 10, len(words))):
                        next_word = words[j]
                        # Check if on the same line (within 8pt vertical)
                        if abs(next_word['top'] - sig_y_plumber) < 8:
                            if '_' in next_word['text']:
                                underline_x0 = next_word['x0']
                                underline_x1 = next_word['x1']
                                found_underline = True
                                print(f"[find_fields] Found text underline after Signature: '{next_word['text']}' x0={underline_x0:.1f} x1={underline_x1:.1f}", file=sys.stderr)
                                break
                    
                    # Method 2: Look for graphic lines near the Signature word
                    if not found_underline:
                        try:
                            lines = last_page.lines or []
                            rects = last_page.rects or []
                            # Check horizontal lines near the signature Y position
                            for line in lines:
                                line_top = line.get('top', 0)
                                line_x0 = line.get('x0', 0)
                                line_x1 = line.get('x1', 0)
                                # Horizontal line near signature, to the right of the word
                                if abs(line_top - sig_y_plumber) < 12 and line_x0 >= sig_label_right - 5 and (line_x1 - line_x0) > 50:
                                    underline_x0 = line_x0
                                    underline_x1 = line_x1
                                    found_underline = True
                                    print(f"[find_fields] Found graphic line after Signature: x0={underline_x0:.1f} x1={underline_x1:.1f} top={line_top:.1f}", file=sys.stderr)
                                    break
                        except Exception as le:
                            print(f"[find_fields] Line detection error: {le}", file=sys.stderr)
                    
                    # Place signature centered on the underline area, slightly above the baseline
                    underline_width = underline_x1 - underline_x0
                    sig_x = underline_x0
                    # Convert to ReportLab Y (from bottom), offset upward so signature sits on the line
                    sig_y = page_height - sig_bottom_plumber + 5
                    field_positions['signature'] = (sig_x, sig_y)
                    field_positions['signature_width'] = max(underline_width, 150)
                    print(f"[find_fields] Found signature at x={sig_x:.1f}, y={sig_y:.1f} width={underline_width:.1f} (plumber_y={sig_y_plumber:.1f})", file=sys.stderr)
            
            # Look for the EXECUTION section labels on the last page
            # These are typically in a table row with Full Name | Date | Place
            for i, word in enumerate(words):
                text = word['text'].lower()
                
                if text == 'full' and i + 1 < len(words) and words[i+1]['text'].lower() == 'name':
                    # "Full Name" label - the value goes below it
                    label_x = word['x0']
                    label_bottom_plumber = word['bottom']
                    # Place text below the label, convert to ReportLab Y
                    field_positions['full_name'] = (label_x + 5, page_height - label_bottom_plumber - 12)
                    print(f"[find_fields] Found full_name label at x={label_x:.1f}, plumber_bottom={label_bottom_plumber:.1f}", file=sys.stderr)
                
                elif text == 'date' and (i == 0 or words[i-1]['text'].lower() != 'trip'):
                    label_x = word['x0']
                    label_bottom_plumber = word['bottom']
                    field_positions['date'] = (label_x + 5, page_height - label_bottom_plumber - 12)
                    print(f"[find_fields] Found date label at x={label_x:.1f}, plumber_bottom={label_bottom_plumber:.1f}", file=sys.stderr)
                
                elif text == 'place':
                    label_x = word['x0']
                    label_bottom_plumber = word['bottom']
                    field_positions['place'] = (label_x + 5, page_height - label_bottom_plumber - 12)
                    print(f"[find_fields] Found place label at x={label_x:.1f}, plumber_bottom={label_bottom_plumber:.1f}", file=sys.stderr)
            
            print(f"[find_fields] Final positions (ReportLab coords): {field_positions}", file=sys.stderr)
            return field_positions if field_positions else None
    except Exception as e:
        import traceback
        print(f"[warn] Could not find PDF fields: {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        return None


def embed_signature_in_pdf(customer, signature_image_path, indemnity_request):
    """Embed customer signature into the indemnity PDF at the signature line.
    
    Fills the EXECUTION table with:
    - Full Name (from KYC data or customer name)
    - Date (signed date)
    - Place (location from signature metadata)
    
    Places signature in the SIGNATURE OF PARTICIPANT box.
    
    Args:
        customer: KYCCustomer object
        signature_image_path: Path to the saved signature image
        indemnity_request: IndemnityRequest object with signature data
    
    Returns:
        Path to the signed PDF with embedded signature
    """
    import os
    from PyPDF2 import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
    from io import BytesIO
    from werkzeug.utils import secure_filename
    
    try:
        # Get the original processed PDF or template PDF
        template = IndemnityTemplate.query.filter_by(id=indemnity_request.template_id).first()
        if not template or not template.pdf_path:
            print("[warn] No template found for signature embedding")
            return None
        
        # Check if there's already a processed PDF for this customer
        base_dir = os.path.dirname(os.path.abspath(__file__))
        processed_dir = os.path.join(UPLOAD_FOLDER, 'kyc', 'processed')
        
        # Look for existing processed PDF for this customer
        existing_processed = None
        if os.path.exists(processed_dir):
            for f in os.listdir(processed_dir):
                if f.startswith(f"processed_{secure_filename(customer.name.replace(' ', '_'))}") and f.endswith('.pdf'):
                    existing_processed = os.path.join(processed_dir, f)
                    break
        
        # Use existing processed PDF or template PDF
        if existing_processed and os.path.exists(existing_processed):
            pdf_path = existing_processed
        else:
            # Resolve template PDF path using UPLOAD_FOLDER
            rel_path = template.pdf_path.replace('uploads/', '', 1) if template.pdf_path.startswith('uploads/') else template.pdf_path
            pdf_path = os.path.join(UPLOAD_FOLDER, rel_path)
            if not os.path.exists(pdf_path):
                # Fallback to base_dir
                pdf_path = os.path.join(base_dir, template.pdf_path)
        
        if not os.path.exists(pdf_path):
            print(f"[warn] PDF not found for signature embedding: {pdf_path}")
            return None
        
        # Read the PDF
        reader = PdfReader(pdf_path)
        writer = PdfWriter()
        
        # Get page size
        first_page = reader.pages[0]
        page_width = float(first_page.mediabox.width)
        page_height = float(first_page.mediabox.height)
        
        # Parse signature data for location
        location = ""
        if indemnity_request.terms_accepted_location:
            location = indemnity_request.terms_accepted_location
        elif indemnity_request.signature_data:
            # Try to extract location from signature_data string
            for part in indemnity_request.signature_data.split('|'):
                if part.startswith('loc:'):
                    location = part.replace('loc:', '')
                    break
        
        # Get full name from KYC submission if available
        full_name = customer.name
        if customer.submission and customer.submission.form_data:
            import json
            try:
                form_data = json.loads(customer.submission.form_data)
                if form_data.get('full_name'):
                    full_name = form_data['full_name']
            except:
                pass
        
        # Format date
        signed_date = ""
        if indemnity_request.signed_at:
            signed_date = indemnity_request.signed_at.strftime('%d/%m/%Y')
        
        # Try to find field positions dynamically
        import sys
        field_positions = find_pdf_fields(pdf_path) or {}
        print(f"[embed_sig] field_positions: {field_positions}", file=sys.stderr)
        print(f"[embed_sig] page_width={page_width}, page_height={page_height}", file=sys.stderr)
        print(f"[embed_sig] full_name='{full_name}', signed_date='{signed_date}', location='{location}'", file=sys.stderr)
        print(f"[embed_sig] signature_image_path='{signature_image_path}'", file=sys.stderr)
        
        # Process each page
        for page_num, page in enumerate(reader.pages):
            # Create overlay canvas
            packet = BytesIO()
            c = canvas.Canvas(packet, pagesize=(page_width, page_height))
            
            # Add signature and execution details on the last page
            if page_num == len(reader.pages) - 1:
                try:
                    # ===== SIGNATURE OF PARTICIPANT BOX =====
                    # Place signature image/text near the "Signature:" line
                    if signature_image_path:
                        # Resolve signature image path
                        sig_rel = signature_image_path.replace('uploads/', '', 1) if signature_image_path.startswith('uploads/') else signature_image_path
                        full_sig_path = os.path.join(UPLOAD_FOLDER, sig_rel)
                        if not os.path.exists(full_sig_path):
                            # Also check without kyc/ prefix
                            full_sig_path = os.path.join(UPLOAD_FOLDER, 'kyc', 'signatures', os.path.basename(sig_rel))
                        if not os.path.exists(full_sig_path):
                            full_sig_path = os.path.join(base_dir, signature_image_path) if not signature_image_path.startswith('/') else signature_image_path
                        
                        print(f"[embed_sig] Resolved sig path: '{full_sig_path}', exists: {os.path.exists(full_sig_path)}", file=sys.stderr)
                        
                        if 'signature' in field_positions:
                            sig_x, sig_y = field_positions['signature']
                        else:
                            # Fallback: place signature in the "SIGNATURE OF PARTICIPANT" box area
                            sig_x = 200
                            sig_y = 155
                        
                        # Use detected underline width or default
                        sig_draw_width = field_positions.get('signature_width', 150)
                        sig_draw_height = int(sig_draw_width * 0.3)  # maintain aspect ratio
                        
                        if signature_image_path.startswith('typed:'):
                            sig_text = signature_image_path.replace('typed:', '')
                            c.setFont("Times-Italic", 18)
                            c.setFillColorRGB(0, 0, 0.8)
                            c.drawString(sig_x, sig_y, sig_text)
                        elif os.path.exists(full_sig_path):
                            img = ImageReader(full_sig_path)
                            c.drawImage(img, sig_x, sig_y, width=sig_draw_width, height=sig_draw_height, mask='auto')
                            print(f"[embed_sig] Drew signature image at ({sig_x}, {sig_y}) size={sig_draw_width}x{sig_draw_height}", file=sys.stderr)
                        else:
                            print(f"[embed_sig] Signature file not found: {full_sig_path}", file=sys.stderr)
                    
                    # ===== NAME, DATE, PLACE below the signature box =====
                    # Since the PDF has no dedicated placeholders for these fields,
                    # we place them in the white space below the signature box
                    c.setFont("Helvetica-Bold", 9)
                    c.setFillColorRGB(0, 0, 0)
                    
                    if field_positions.get('full_name'):
                        x, y = field_positions['full_name']
                        c.drawString(x, y, full_name)
                    elif field_positions.get('signature'):
                        # Place below signature
                        sx, sy = field_positions['signature']
                        c.drawString(sx - 100, sy - 55, f"Full Name: {full_name}")
                        c.setFont("Helvetica", 9)
                        c.drawString(sx - 100, sy - 70, f"Date: {signed_date}")
                        c.drawString(sx - 100, sy - 85, f"Place: {location}")
                    else:
                        # Absolute fallback - place in the white space below signature area
                        info_y = 95
                        c.drawString(100, info_y, f"Full Name: {full_name}")
                        c.setFont("Helvetica", 9)
                        c.drawString(100, info_y - 15, f"Date: {signed_date}")
                        c.drawString(100, info_y - 30, f"Place: {location}")
                    
                    # Add execution details for date/place if separate positions found
                    if field_positions.get('date'):
                        c.setFont("Helvetica-Bold", 9)
                        x, y = field_positions['date']
                        c.drawString(x, y, signed_date)
                    
                    if field_positions.get('place'):
                        c.setFont("Helvetica-Bold", 9)
                        x, y = field_positions['place']
                        c.drawString(x, y, location)
                    
                    # Add timestamp in smaller text
                    c.setFont("Helvetica", 7)
                    c.setFillColorRGB(0.4, 0.4, 0.4)
                    
                    from datetime import timedelta
                    if indemnity_request.signed_at:
                        ist_offset = timedelta(hours=5, minutes=30)
                        ist_time = (indemnity_request.signed_at + ist_offset).strftime('%d-%m-%Y %I:%M %p')
                    else:
                        ist_time = ''
                    
                    timestamp_y = 60
                    if field_positions.get('signature'):
                        sx, sy = field_positions['signature']
                        timestamp_y = sy - 100
                    
                    c.drawString(100, timestamp_y, f"Digitally signed on {ist_time} IST | IP: {indemnity_request.ip_address or 'N/A'}")
                            
                except Exception as sig_e:
                    import traceback
                    print(f"[warn] Failed to embed signature elements: {sig_e}", file=sys.stderr)
                    traceback.print_exc(file=sys.stderr)
            
            c.save()
            packet.seek(0)
            
            # Merge overlay with page
            try:
                from PyPDF2 import PdfReader as OverlayReader
                overlay = OverlayReader(packet)
                if len(overlay.pages) > 0:
                    page.merge_page(overlay.pages[0])
            except Exception as e:
                print(f"[warn] Could not merge signature overlay: {e}")
            
            writer.add_page(page)
        
        # Create signed PDF filename
        signed_filename = f"signed_{secure_filename(customer.name.replace(' ', '_'))}_{generate_token(8)}.pdf"
        signed_dir = os.path.join(UPLOAD_FOLDER, 'kyc', 'signed')
        os.makedirs(signed_dir, exist_ok=True)
        signed_path = os.path.join(signed_dir, signed_filename)
        
        # Write the signed PDF
        with open(signed_path, 'wb') as output_file:
            writer.write(output_file)
        
        # Return relative path
        signed_relative_path = os.path.join('uploads', 'kyc', 'signed', signed_filename)
        print(f"[info] Created signed PDF: {signed_relative_path}")
        return signed_relative_path
        
    except Exception as e:
        print(f"[error] Failed to embed signature in PDF: {e}")
        import traceback
        traceback.print_exc()
        return None


@app.route('/kyc/submissions/<int:customer_id>')
@login_required
def kyc_submission_view(customer_id):
    """View a specific customer's KYC submission (for admin/agents)."""
    import json
    customer = KYCCustomer.query.get_or_404(customer_id)
    
    # Check permission
    if current_user.role != 'admin' and customer.created_by_id != current_user.id:
        flash('You do not have permission to view this submission.', 'danger')
        return redirect(url_for('kyc_customers'))
    
    form_data = {}
    document_paths = {}
    
    if customer.submission:
        form_data = json.loads(customer.submission.form_data) if customer.submission.form_data else {}
        document_paths = json.loads(customer.submission.document_paths) if customer.submission.document_paths else {}
    
    return render_template('kyc/submission_view.html',
                         customer=customer,
                         form_data=form_data,
                         document_paths=document_paths)


@app.route('/kyc/customers/<int:customer_id>/detail')
@login_required
def kyc_customer_detail(customer_id):
    """View detailed customer information with KYC data and documents."""
    import json
    customer = KYCCustomer.query.get_or_404(customer_id)
    
    # Check permission
    if current_user.role != 'admin' and customer.created_by_id != current_user.id:
        flash('You do not have permission to view this customer.', 'danger')
        return redirect(url_for('kyc_customers'))
    
    # Get form data and document paths
    form_data = {}
    document_paths = {}
    indemnity_request = None
    
    if customer.submission:
        form_data = json.loads(customer.submission.form_data) if customer.submission.form_data else {}
        document_paths = json.loads(customer.submission.document_paths) if customer.submission.document_paths else {}
    
    # Phase 6c: load the SIGNED indemnity request (if any) using derived
    # status so stuck-flag customers still see their signed PDF. We pick
    # the signed row explicitly (a customer may have multiple requests if
    # the email was resent before signature).
    if customer.effective_indemnity_signed:
        indemnity_request = (
            IndemnityRequest.query
            .filter_by(customer_id=customer.id)
            .filter(IndemnityRequest.signed_at.isnot(None))
            .order_by(IndemnityRequest.signed_at.desc())
            .first()
        )
        # Fallback: no signed row yet visible (rare race) — use latest request
        # so the template still has a reference to render metadata.
        if indemnity_request is None:
            indemnity_request = (
                IndemnityRequest.query
                .filter_by(customer_id=customer.id)
                .order_by(IndemnityRequest.sent_at.desc())
                .first()
            )

    # Get active template for indemnity PDF
    active_template = IndemnityTemplate.query.filter_by(is_active=True).first()

    # Phase 4: previous trip count (read-only, best-effort)
    previous_trip_count = customer.count_previous_trips()

    return render_template('kyc/customer_detail.html',
                         customer=customer,
                         form_data=form_data,
                         document_paths=document_paths,
                         indemnity_request=indemnity_request,
                         active_template=active_template,
                         previous_trip_count=previous_trip_count)


@app.route('/kyc/settings/email', methods=['GET', 'POST'])
@login_required
def kyc_email_settings():
    """Configure SMTP settings for KYC emails (care@deyor.in)."""
    if request.method == 'POST':
        smtp_host = request.form.get('smtp_host', '').strip()
        smtp_port = request.form.get('smtp_port', '587').strip()
        smtp_user = request.form.get('smtp_user', '').strip()
        smtp_pass = request.form.get('smtp_pass', '').strip()
        
        if not smtp_host or not smtp_user or not smtp_pass:
            flash('SMTP Host, Username, and Password are required.', 'danger')
            return redirect(url_for('kyc_email_settings'))
        
        # Save settings to database
        settings = [
            ('kyc_smtp_host', smtp_host),
            ('kyc_smtp_port', smtp_port),
            ('kyc_smtp_user', smtp_user),
            ('kyc_smtp_pass', smtp_pass),
        ]
        
        for key, value in settings:
            meta = AppMeta.query.filter_by(key=key).first()
            if meta:
                meta.value = value
            else:
                meta = AppMeta(key=key, value=value)
                db.session.add(meta)
        
        db.session.commit()
        flash('Email settings saved successfully.', 'success')
        return redirect(url_for('kyc_email_settings'))
    
    # Get current settings
    settings = get_kyc_email_settings()
    return render_template('kyc/email_settings.html', settings=settings)


@app.route('/kyc/settings/email/test', methods=['POST'])
@login_required
def kyc_test_email():
    """Test email configuration by sending a test email."""
    test_email = request.form.get('test_email', current_user.email).strip()
    
    try:
        settings = get_kyc_email_settings()
        
        if not settings['smtp_host'] or not settings['smtp_user'] or not settings['smtp_pass']:
            flash('Please save SMTP settings first.', 'danger')
            return redirect(url_for('kyc_email_settings'))
        
        subject = "Test Email from Deyor KYC System"
        body = f"""Hello,

This is a test email from your Deyor KYC email configuration.

If you received this email, your SMTP settings are working correctly!

Configuration used:
- From: {settings['smtp_user']}
- SMTP Host: {settings['smtp_host']}
- SMTP Port: {settings['smtp_port']}

Best regards,
Deyor System
"""
        
        context = ssl.create_default_context(cafile=certifi.where())
        with smtplib.SMTP(settings['smtp_host'], settings['smtp_port']) as server:
            server.starttls(context=context)
            server.login(settings['smtp_user'], settings['smtp_pass'])
            
            msg = MIMEText(body)
            msg['Subject'] = subject
            msg['From'] = settings['smtp_user']
            msg['To'] = test_email
            
            server.sendmail(settings['smtp_user'], [test_email], msg.as_string())
        
        flash(f'Test email sent successfully to {test_email}!', 'success')
    except Exception as e:
        flash(f'Failed to send test email: {str(e)}', 'danger')
    
    return redirect(url_for('kyc_email_settings'))


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """Serve uploaded files securely from Render disk or local uploads."""
    from flask import send_from_directory, abort, current_app, jsonify
    import os
    
    # Debug logging - print to stderr for Render logs
    import sys
    print(f"[UPLOADS] Requested filename: '{filename}'", file=sys.stderr)
    print(f"[UPLOADS] UPLOAD_FOLDER: '{UPLOAD_FOLDER}'", file=sys.stderr)
    print(f"[UPLOADS] UPLOAD_PATH env: '{os.environ.get('UPLOAD_PATH')}'", file=sys.stderr)
    
    # The filename from URL is like 'kyc/indemnity_xxx.pdf'
    # We need to look in UPLOAD_FOLDER for this path
    
    # Try direct path first
    full_path = os.path.join(UPLOAD_FOLDER, filename)
    print(f"[UPLOADS] Checking full_path: '{full_path}' exists: {os.path.exists(full_path)}", file=sys.stderr)
    
    if os.path.exists(full_path):
        print(f"[UPLOADS] Serving from UPLOAD_FOLDER", file=sys.stderr)
        return send_from_directory(UPLOAD_FOLDER, filename)
    
    # Try without 'kyc/' prefix if filename starts with it
    if filename.startswith('kyc/'):
        alt_filename = filename[4:]  # Remove 'kyc/'
        alt_path = os.path.join(UPLOAD_FOLDER, 'kyc', alt_filename)
        print(f"[UPLOADS] Checking alt_path: '{alt_path}' exists: {os.path.exists(alt_path)}", file=sys.stderr)
        if os.path.exists(alt_path):
            return send_from_directory(os.path.join(UPLOAD_FOLDER, 'kyc'), alt_filename)
    
    # Try with 'kyc/' prefix if filename doesn't have it
    if not filename.startswith('kyc/'):
        alt_path = os.path.join(UPLOAD_FOLDER, 'kyc', filename)
        print(f"[UPLOADS] Checking alt_path with kyc/: '{alt_path}' exists: {os.path.exists(alt_path)}", file=sys.stderr)
        if os.path.exists(alt_path):
            return send_from_directory(os.path.join(UPLOAD_FOLDER, 'kyc'), filename)
    
    # Handle path mismatch: files saved to UPLOAD_FOLDER/subdir/ but URL expects kyc/subdir/
    # e.g., file at /uploads/signatures/sig.png but URL says kyc/signatures/sig.png
    if filename.startswith('kyc/'):
        stripped = filename[4:]  # Remove 'kyc/' prefix
        alt_path = os.path.join(UPLOAD_FOLDER, stripped)
        print(f"[UPLOADS] Checking without kyc/ prefix: '{alt_path}' exists: {os.path.exists(alt_path)}", file=sys.stderr)
        if os.path.exists(alt_path):
            return send_from_directory(os.path.dirname(alt_path), os.path.basename(alt_path))
    
    # Try just the basename in various subdirectories
    basename = os.path.basename(filename)
    for subdir in ['kyc', 'kyc/signatures', 'kyc/signed', 'kyc/processed', 'signatures', 'signed', 'processed']:
        alt_path = os.path.join(UPLOAD_FOLDER, subdir, basename)
        if os.path.exists(alt_path):
            print(f"[UPLOADS] Found via basename search: '{alt_path}'", file=sys.stderr)
            return send_from_directory(os.path.join(UPLOAD_FOLDER, subdir), basename)
    
    print(f"[UPLOADS] File not found: '{filename}'", file=sys.stderr)
    abort(404)


def _commit_and_verify_kyc_status(customer, flag_field):
    """Commit pending KYC status change and verify it actually persisted.

    This exists because we have observed intermittent cases where the
    success email is sent but the customer's `kyc_submitted` /
    `indemnity_signed` flag remains False in the dashboard. The root
    causes can be: a silently-rolled-back transaction, a stale connection
    from the pool, or a race where the commit appeared to succeed in
    memory but did not persist.

    This helper:
      1. Attempts `db.session.commit()` and rolls back on any exception.
      2. Forces a fresh read from the database via `session.expire()`
         so we are not reading the in-memory attribute.
      3. Returns (True, None) only if the flag is truly True in the DB.
      4. NEVER deletes or mutates any other row. It only verifies.

    Returns:
        (success: bool, error_message: str | None)
    """
    import sys
    customer_id = getattr(customer, 'id', None)
    try:
        db.session.commit()
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        msg = f"[kyc-commit] DB commit FAILED customer_id={customer_id} flag={flag_field}: {e}"
        try:
            app.logger.error(msg)
        except Exception:
            print(msg, file=sys.stderr)
        return False, "We could not save your submission due to a database error. Please try again."

    # Phase 6c: Force a fresh read with a short retry loop. Postgres MVCC
    # can occasionally show a stale snapshot on the first read after a
    # commit when the connection is reused from the pool. Two extra
    # verification attempts (with a 50 ms sleep) absorb that lag without
    # degrading the happy-path UX.
    import time
    persisted = False
    last_err = None
    for attempt in range(3):
        try:
            db.session.expire(customer)
            persisted = bool(getattr(customer, flag_field, False))
            if persisted:
                break
        except Exception as e:
            last_err = e
            msg = (f"[kyc-commit] Verification read attempt {attempt+1} failed "
                   f"customer_id={customer_id} flag={flag_field}: {e}")
            try:
                app.logger.warning(msg)
            except Exception:
                print(msg, file=sys.stderr)
        # Not the last attempt: small backoff before retrying.
        if attempt < 2:
            time.sleep(0.05)

    if not persisted:
        # Final fallback: consult the source-of-truth row directly. If the
        # source row proves the event happened (KYCSubmission exists for
        # `kyc_submitted`, or a signed IndemnityRequest for
        # `indemnity_signed`), we treat the commit as verified. This
        # eliminates false-negative retries where the flag genuinely was
        # persisted but the re-read missed it. Still read-only.
        try:
            if flag_field == 'kyc_submitted':
                proof = db.session.query(KYCSubmission.id).filter_by(
                    customer_id=customer_id
                ).first() is not None
            elif flag_field == 'indemnity_signed':
                proof = db.session.query(IndemnityRequest.id).filter_by(
                    customer_id=customer_id
                ).filter(IndemnityRequest.signed_at.isnot(None)).first() is not None
            else:
                proof = False
        except Exception as e:
            proof = False
            last_err = e

        if proof:
            info = (f"[kyc-commit] VERIFIED-BY-SOURCE customer_id={customer_id} "
                    f"flag={flag_field} (raw flag was stale, source row present)")
            try:
                app.logger.info(info)
            except Exception:
                print(info, file=sys.stderr)
            return True, None

        msg = (f"[kyc-commit] VERIFICATION MISMATCH customer_id={customer_id} "
               f"flag={flag_field} persisted_value={persisted} last_err={last_err}")
        try:
            app.logger.error(msg)
        except Exception:
            print(msg, file=sys.stderr)
        return False, "Your submission did not save correctly. Please try again."

    info = f"[kyc-commit] VERIFIED customer_id={customer_id} flag={flag_field}=True"
    try:
        app.logger.info(info)
    except Exception:
        print(info, file=sys.stderr)
    return True, None


def create_kyc_notifications(customer, event_type):
    """Create in-app notifications and send email alerts for KYC events.
    
    Args:
        customer: KYCCustomer object
        event_type: 'kyc_completed' or 'indemnity_signed'
    """
    # Get the user who created this customer
    creator = db.session.get(User, customer.created_by_id)
    if not creator:
        return
    
    # Create in-app notification
    if event_type == 'kyc_completed':
        message = f"Customer {customer.name} has completed their KYC form."
        notification_type = 'kyc_completed'
    elif event_type == 'indemnity_signed':
        message = f"Customer {customer.name} has signed the Terms & Conditions and Indemnity Agreement."
        notification_type = 'indemnity_signed'
    else:
        message = f"Update from customer {customer.name}."
        notification_type = 'kyc_update'
    
    # Create notification for the creator
    notification = Notification(
        user_id=creator.id,
        type=notification_type,
        message=message,
        created_at=datetime.utcnow(),
        is_read=False
    )
    db.session.add(notification)
    db.session.commit()
    
    # Send email notification to creator only
    try:
        send_kyc_admin_notification(creator.email, customer, event_type)
    except Exception as e:
        print(f"[warn] Failed to send admin notification email: {e}")


def send_kyc_admin_notification(admin_email, customer, event_type):
    """Send email notification to admin when customer completes KYC action."""
    settings = get_kyc_email_settings()
    
    if not settings['smtp_host'] or not settings['smtp_user'] or not settings['smtp_pass']:
        print("[warn] SMTP not configured, skipping admin notification email")
        return
    
    if event_type == 'kyc_completed':
        subject = f"KYC Completed - {customer.name}"
        body = f"""Hello,

Customer {customer.name} has successfully completed their KYC form.

Customer Details:
- Name: {customer.name}
- Email: {customer.email}
- Phone: {customer.phone}
- Trip Type: {customer.trip_type.title()}
- Trip Date: {customer.trip_date.strftime('%d-%m-%Y') if customer.trip_date else 'Not specified'}
- Booking ID: {customer.booking_id or 'N/A'}

You can view their submission at: {url_for('kyc_customer_detail', customer_id=customer.id, _external=True)}

Best regards,
Deyor KYC System
"""
    elif event_type == 'indemnity_signed':
        subject = f"Indemnity Signed - {customer.name}"
        body = f"""Hello,

Customer {customer.name} has signed the Terms & Conditions and Indemnity Agreement.

Customer Details:
- Name: {customer.name}
- Email: {customer.email}
- Phone: {customer.phone}
- Trip Type: {customer.trip_type.title()}
- Trip Date: {customer.trip_date.strftime('%d-%m-%Y') if customer.trip_date else 'Not specified'}
- Booking ID: {customer.booking_id or 'N/A'}

Signed at: {datetime.utcnow().strftime('%d-%m-%Y %I:%M %p')} IST

You can view their details at: {url_for('kyc_customer_detail', customer_id=customer.id, _external=True)}

Best regards,
Deyor KYC System
"""
    else:
        return
    
    try:
        context = ssl.create_default_context(cafile=certifi.where())
        with smtplib.SMTP(settings['smtp_host'], settings['smtp_port']) as server:
            server.starttls(context=context)
            server.login(settings['smtp_user'], settings['smtp_pass'])
            
            msg = MIMEText(body)
            msg['Subject'] = subject
            msg['From'] = settings['smtp_user']
            msg['To'] = admin_email
            
            server.sendmail(settings['smtp_user'], [admin_email], msg.as_string())
            print(f"[info] Admin notification sent to {admin_email}")
    except Exception as e:
        print(f"[warn] Failed to send admin notification: {e}")


@app.route('/kyc/admin/reprocess-pdfs', methods=['POST'])
@login_required
def kyc_reprocess_pdfs():
    """Admin route to reprocess all signed PDFs with corrected coordinates."""
    if current_user.role != 'admin':
        flash('Only administrators can perform this action.', 'danger')
        return redirect(url_for('kyc_dashboard'))
    
    try:
        # Import the reprocess function
        import subprocess
        import sys
        
        result = subprocess.run([sys.executable, 'reprocess_signed_pdfs.py'], 
                              capture_output=True, text=True, cwd=os.path.dirname(os.path.abspath(__file__)))
        
        if result.returncode == 0:
            flash('Successfully reprocessed all signed PDFs with corrected coordinates.', 'success')
        else:
            flash(f'Reprocessing completed with some issues. Check logs.', 'warning')
            print(f"[reprocess output] {result.stdout}")
            print(f"[reprocess errors] {result.stderr}")
    except Exception as e:
        flash(f'Failed to reprocess PDFs: {str(e)}', 'danger')
    
    return redirect(url_for('kyc_dashboard'))


@app.route('/kyc/debug/indemnity-pdf')
def kyc_debug_indemnity_pdf():
    """Directly serve the indemnity PDF for testing."""
    import os
    from flask import send_from_directory, abort, jsonify
    
    # Get active template
    template = IndemnityTemplate.query.filter_by(is_active=True).first()
    if not template or not template.pdf_path:
        return jsonify({'error': 'No active template or pdf_path'}), 404
    
    # Get the filename from the stored path
    pdf_path = template.pdf_path  # e.g., uploads/kyc/indemnity_xxx.pdf
    filename = pdf_path.replace('uploads/', '')  # e.g., kyc/indemnity_xxx.pdf
    
    # Try to serve directly from UPLOAD_FOLDER
    full_path = os.path.join(UPLOAD_FOLDER, filename)
    
    if os.path.exists(full_path):
        return send_from_directory(UPLOAD_FOLDER, filename)
    
    # Try alternate locations
    possible_paths = [
        os.path.join('/opt/render/project/src/uploads', filename),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', filename),
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            dir_name = os.path.dirname(path)
            base_name = os.path.basename(path)
            return send_from_directory(dir_name, base_name)
    
    return jsonify({
        'error': 'PDF not found',
        'pdf_path': pdf_path,
        'filename': filename,
        'checked_paths': [full_path] + possible_paths,
        'UPLOAD_FOLDER': UPLOAD_FOLDER
    }), 404

@app.route('/kyc/debug/status')
def kyc_debug_status():
    """Debug endpoint to check KYC system status and file storage."""
    import os
    from flask import jsonify
    
    try:
        # Check KYC Form
        kyc_form = KYCForm.query.filter_by(is_active=True).first()
        
        # Check Indemnity Template
        indemnity_template = IndemnityTemplate.query.filter_by(is_active=True).first()
        
        # Check file paths
        render_upload_dir = os.environ.get('UPLOAD_PATH') or '/opt/render/project/src/uploads'
        local_upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
        
        # List files in render disk
        render_files = []
        try:
            if os.path.exists(render_upload_dir):
                for f in os.listdir(os.path.join(render_upload_dir, 'kyc')) if os.path.exists(os.path.join(render_upload_dir, 'kyc')) else []:
                    render_files.append(f)
        except Exception as e:
            render_files = [f"Error: {str(e)}"]
        
        return jsonify({
            'kyc_form_exists': kyc_form is not None,
            'indemnity_template_exists': indemnity_template is not None,
            'indemnity_terms_pdf_path': indemnity_template.terms_pdf_path if indemnity_template else None,
            'indemnity_pdf_path': indemnity_template.pdf_path if indemnity_template else None,
            'storage': {
                'UPLOAD_FOLDER': UPLOAD_FOLDER,
                'UPLOAD_PATH_env': os.environ.get('UPLOAD_PATH'),
                'render_dir': render_upload_dir,
                'render_dir_exists': os.path.exists(render_upload_dir),
                'kyc_dir_exists': os.path.exists(os.path.join(render_upload_dir, 'kyc')),
                'files_in_kyc': render_files[:10]
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.getenv('PORT', '5000'))
    app.run(host='0.0.0.0', port=port, debug=True)
